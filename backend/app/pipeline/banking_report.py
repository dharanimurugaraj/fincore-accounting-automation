import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from .engine import FinCoreComputationEngine

def generate_banking_report(
    accounts_data: list,
    computed: dict,
    job_id: str,
    period: str,
    loan_results: list = None,  # Gap 6: list of dicts from calculate_loan_interest()
) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # 1. Detail Sheet
    detail_ws = wb.create_sheet("Detail Sheet")
    
    # Header logic
    bank_name = accounts_data[0].get("bank_name", "UNKNOWN").upper() if accounts_data else "UNKNOWN"
    acct_nos = [str(a.get("account_number", "UNKNOWN"))[-3:] for a in accounts_data]
    acct_str = ", ".join(acct_nos)
    
    detail_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(accounts_data) + 7)
    limit_val = accounts_data[0].get("cc_limit") or 0.0
    header_cell = detail_ws.cell(row=1, column=1)
    header_cell.value = f"CONSOLIDATED BANKING REPORT | {bank_name} - {acct_str} | Limit: ₹{limit_val:,.2f} | {period}"
    header_cell.font = Font(bold=True, size=12)
    header_cell.alignment = Alignment(horizontal="center")

    # Aggregate by date
    all_dates = {}
    acct_labels = []
    
    for i, account in enumerate(accounts_data, 1):
        acct_no = str(account.get("account_number", "UNKNOWN"))[-3:]
        label = f"Acct{i}-CC ({acct_no})"
        acct_labels.append(label)
        
        txns = account.get("full_month_transactions", [])
        for txn in txns:
            dt = txn.get("date")
            if dt not in all_dates: all_dates[dt] = {}
            # CC Value logic from plan: if closing balance is negative, CC = abs(bal)
            bal = float(txn.get("closing_balance", 0))
            all_dates[dt][label] = abs(bal) if bal < 0 else 0

    # Column headers
    headers = ["Date"] + acct_labels + ["CC Drawing", "WCDL", "Total Utilisation", "ROI", "CC Int. Rate (/day)"]
    header_fill = PatternFill(fill_type="solid", fgColor="0A0A0F")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    for col, h in enumerate(headers, 1):
        cell = detail_ws.cell(row=3, column=col)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    sorted_dates = sorted(all_dates.keys())
    # Detail Column Indices
    cc_util_col = len(acct_labels) + 2
    wcdl_col = len(acct_labels) + 3
    total_util_col = len(acct_labels) + 4
    roi_col = len(acct_labels) + 5
    int_rate_col = len(acct_labels) + 6

    # Accumulate data for snapshot sheet
    cumulative_interest = 0
    period_start = sorted_dates[0] if sorted_dates else period
    period_end = sorted_dates[-1] if sorted_dates else period
    period_label = f"{period_start} to {period_end}"

    for i, dt in enumerate(sorted_dates):
        row_idx = i + 4
        
        # [Date]
        date_cell = detail_ws.cell(row=row_idx, column=1)
        date_cell.value = dt
        date_cell.border = thin_border
        
        # [Accounts]
        for j, label in enumerate(acct_labels):
            cell = detail_ws.cell(row=row_idx, column=j + 2)
            cell.value = all_dates[dt].get(label, 0)
            cell.number_format = '#,##0.00'
            cell.border = thin_border
            
        # [CC Utilisation]
        start_col = openpyxl.utils.get_column_letter(2)
        end_col = openpyxl.utils.get_column_letter(len(acct_labels) + 1)
        util_cell = detail_ws.cell(row=row_idx, column=cc_util_col)
        util_cell.value = f"=SUM({start_col}{row_idx}:{end_col}{row_idx})"
        util_cell.number_format = '#,##0.00'
        util_cell.border = thin_border
        
        # [WCDL]
        # For consolidated report, WCDL defaults to 0 unless specifically provided per day
        wcdl_ref_cell = detail_ws.cell(row=row_idx, column=wcdl_col)
        wcdl_ref_cell.value = 0.0
        wcdl_ref_cell.number_format = '#,##0.00'
        wcdl_ref_cell.border = thin_border
        
        # [Total Utilisation]
        cc_util_letter = get_column_letter(cc_util_col)
        wcdl_letter = get_column_letter(wcdl_col)
        total_cell = detail_ws.cell(row=row_idx, column=total_util_col)
        total_cell.value = f"={cc_util_letter}{row_idx}+{wcdl_letter}{row_idx}"
        total_cell.number_format = '#,##0.00'
        total_cell.border = thin_border
        
        # [ROI]
        roi_val = (accounts_data[0].get("cc_roi_percent") or 7.60) / 100
        roi_cell = detail_ws.cell(row=row_idx, column=roi_col)
        roi_cell.value = roi_val
        roi_cell.number_format = '0.00%'
        roi_cell.border = thin_border
        
        # [CC Int Rate]
        total_util_letter = get_column_letter(total_util_col)
        roi_letter = get_column_letter(roi_col)
        int_cell = detail_ws.cell(row=row_idx, column=int_rate_col)
        int_cell.value = f"={total_util_letter}{row_idx}*{roi_letter}{row_idx}/365"
        int_cell.number_format = '#,##0.00'
        int_cell.border = thin_border
        
        # Snapshot accumulation
        day_util = sum([all_dates[dt].get(l, 0) for l in acct_labels])
        cumulative_interest += (day_util * (roi_val) / 365)

    # Set Column Widths for readability
    detail_ws.column_dimensions['A'].width = 15
    for i in range(2, int_rate_col + 1):
        col_letter = get_column_letter(i)
        detail_ws.column_dimensions[col_letter].width = 22

    # 2. Snapshot Sheet (Section A/B/C)
    snapshot_ws = wb.create_sheet("Report Snapshot", 0)
    _populate_snapshot(snapshot_ws, accounts_data, computed, period_label, loan_results or [])

    # Save file — use /tmp on Vercel (read-only filesystem), local path otherwise
    _is_vercel = os.getenv("VERCEL") or os.getenv("AWS_LAMBDA_FUNCTION_NAME")
    storage_dir = "/tmp/fincore/banking_reports" if _is_vercel else "./storage/banking_reports"
    os.makedirs(storage_dir, exist_ok=True)
    date_prefix = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{date_prefix}_bankingreport.xlsx"
    filepath = os.path.join(storage_dir, filename)
    wb.save(filepath)

    return filepath

def _populate_snapshot(ws, accounts_data, computed, period_label, loan_results=None):
    """
    Builds the high-level report snapshot — Section A, B, and C.

    Gap 6: Adds blended_roi_pct, finance_cost_pct display, and per-loan ROI
           verification table (Section C) sourced from loan_results.
    """
    loan_results = loan_results or []
    header_fill   = PatternFill(fill_type="solid", fgColor="E0E0E0")
    section_fill  = PatternFill(fill_type="solid", fgColor="0A0A0F")
    section_font  = Font(bold=True, color="FFFFFF")
    match_fill    = PatternFill(fill_type="solid", fgColor="D4EDDA")
    flag_fill     = PatternFill(fill_type="solid", fgColor="FFF3CD")

    ws.cell(1, 1).value = f"Banking Report — {period_label}"
    ws.cell(1, 1).font  = Font(size=14, bold=True)

    # ── Section A: Working Capital Utilisation ────────────────────────────────
    ws.cell(3, 1).value = "SECTION A — Working Capital Funds: Utilisation"
    ws.cell(3, 1).font  = Font(bold=True)
    ws.cell(3, 1).fill  = section_fill
    ws.cell(3, 1).font  = section_font
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)

    a_headers = ["Bank", "Sanctioned (Cr)", "Avg CC (Cr)", "Avg WCDL (Cr)", "Total Avg (Cr)", "% Utilisation"]
    for col, h in enumerate(a_headers, 1):
        cell = ws.cell(5, col)
        cell.value = h
        cell.font  = Font(bold=True)
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal="center")

    INR_CR = 10_000_000  # 1 Crore
    total_sanct_limit = max([(a.get("cc_limit") or 0) for a in accounts_data], default=0)
    limit_display     = round(total_sanct_limit / INR_CR, 2) if total_sanct_limit > 0 else "N/A"
    total_avg_cc_util = computed.get("average_utilisation", 0) / INR_CR
    wcdl_avg_util     = computed.get("avg_wcdl_util", 0) / INR_CR
    grand_total_util  = total_avg_cc_util + wcdl_avg_util
    util_pct          = (
        (grand_total_util * INR_CR) / total_sanct_limit * 100
        if total_sanct_limit > 0 else 0
    )
    bank_name = accounts_data[0].get("bank_name", "UNKNOWN").upper() if accounts_data else "UNKNOWN"
    for col, val in enumerate(
        [bank_name, limit_display, round(total_avg_cc_util, 2),
         round(wcdl_avg_util, 2), round(grand_total_util, 2), f"{util_pct:.2f}%"],
        1,
    ):
        ws.cell(6, col).value = val

    # ── Section A right panel: Finance Cost ───────────────────────────────────
    ws.cell(8, 1).value = "SECTION A — Finance Cost"
    ws.cell(8, 1).fill  = section_fill
    ws.cell(8, 1).font  = section_font
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=3)

    total_int      = computed.get("total_interest", 0)
    blended_roi    = computed.get("blended_roi_pct", 0)
    finance_cost   = computed.get("finance_cost_pct", 0)
    days_in_period = computed.get("days_in_period", 30)

    finance_kpis = [
        ("Interest Charged — CC",            f"₹{computed.get('total_cc_interest', 0):,.2f}"),
        ("Interest Charged — WCDL/BC/PQL",   f"₹{computed.get('total_wcdl_interest', 0):,.2f}"),
        ("Total Interest Charged",            f"₹{total_int:,.2f}"),
        ("Avg Utilisation (Denominator INR)", f"₹{computed.get('average_utilisation', 0):,.2f}"),
        ("Days in Period",                    str(days_in_period)),
        ("Blended Interest Cost % (ann.)",    f"{blended_roi:.4f}%"),   # Gap 6
        ("Finance Cost % (ann.)",             f"{finance_cost:.4f}%"),   # Gap 6
        ("Notional Interest Loss (Cr bal)",   f"₹{computed.get('total_notional_loss', 0):,.2f}"),
    ]
    for idx, (label, value) in enumerate(finance_kpis, 9):
        ws.cell(idx, 1).value = label
        ws.cell(idx, 2).value = value
        ws.cell(idx, 1).font  = Font(bold=True)

    # ── Section B: Bank Charges ───────────────────────────────────────────────
    b_row = 9 + len(finance_kpis) + 1
    ws.cell(b_row, 1).value = "SECTION B — Bank Charges"
    ws.cell(b_row, 1).fill  = section_fill
    ws.cell(b_row, 1).font  = section_font
    ws.merge_cells(start_row=b_row, start_column=1, end_row=b_row, end_column=3)
    ws.cell(b_row + 1, 1).value = "Total Bank Charges"
    ws.cell(b_row + 1, 2).value = "Refer Working Sheet"

    # ── Section C: ROI Verification per Loan (Gap 6) ──────────────────────────
    c_row = b_row + 4
    ws.cell(c_row, 1).value = "SECTION C — Per-Loan ROI Verification (WCDL / BC / PQL)"
    ws.cell(c_row, 1).fill  = section_fill
    ws.cell(c_row, 1).font  = section_font
    ws.merge_cells(start_row=c_row, start_column=1, end_row=c_row, end_column=9)

    c_headers = [
        "Loan Ref", "Type", "Bank",
        "Principal (INR)", "ROI %",
        "Drawdown", "Maturity / Prepay",
        "Actual Days", "Our Calc (₹)",
        "Bank Charged (₹)", "Diff (₹)", "Status",
    ]
    for col, h in enumerate(c_headers, 1):
        cell = ws.cell(c_row + 1, col)
        cell.value = h
        cell.font  = Font(bold=True)
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r_i, loan in enumerate(loan_results, c_row + 2):
        row_vals = [
            loan.get("loan_ref", ""),
            loan.get("loan_type", ""),
            loan.get("bank_name", ""),
            loan.get("principal_inr", 0),
            f"{float(loan.get('loan_roi', 0)) * 100:.4f}%",
            loan.get("drawdown_date", ""),
            loan.get("prepayment_date") or loan.get("maturity_date", ""),
            loan.get("actual_days", ""),
            loan.get("calculated_interest", 0),
            loan.get("interest_as_per_bank", "Pending"),
            loan.get("diff_vs_bank", ""),
            loan.get("diff_status", "PENDING"),
        ]
        status = loan.get("diff_status", "PENDING")
        row_fill = match_fill if status == "✓ MATCH" else (flag_fill if status == "⚠ FLAG" else None)

        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(r_i, col)
            cell.value = val
            if row_fill:
                cell.fill = row_fill
            if col in (4, 9, 10, 11):  # numeric cols
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    for c_i in range(3, 13):
        ws.column_dimensions[get_column_letter(c_i)].width = 18
