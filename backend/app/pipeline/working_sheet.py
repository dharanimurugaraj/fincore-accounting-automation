import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border
from datetime import datetime, timedelta
import os
from typing import Optional, Tuple

from .engine import FinCoreComputationEngine


def _last_row_index_by_date(full_month_txns: list) -> dict:
    """Map YYYY-MM-DD → row index of the *last* line for that date (end-of-day closing)."""
    last_by: dict = {}
    for idx, t in enumerate(full_month_txns):
        dk = (t.get("date") or "")[:10]
        if dk:
            last_by[dk] = idx
    return last_by


def _positive_bal_num_days_cc(
    cb: float,
    balance_style: str,
    balance_credit_side: Optional[bool],
) -> Tuple[float, int, float]:
    """
    Apply User Rule:
    - Positive Bal: If Cr marker exists, use absolute value. Else 0.
    - No. of Days: If Cr marker exists, use 1. Else 0.
    - CC Util: abs(cb) if balance is OD (marker False or signed negative).
    """
    cb = float(cb or 0.0)
    
    # 1. Positive Bal & No. of Days (Strictly tied to Cr marker per User Request)
    if balance_credit_side is True:
        pos_bal = abs(cb)
        num_days = 1
    else:
        # If OD, Dr, or unknown (None), Pos Bal is 0
        pos_bal = 0.0
        num_days = 0

    # 2. CC Utilisation (for interest/drawn calculation)
    if balance_credit_side is False:
        cc_amt = abs(cb)
    elif balance_credit_side is True:
        cc_amt = 0.0
    else:
        # Fallback for plain signed numbers where positive = credit
        cc_amt = abs(cb) if cb < 0 else 0.0

    return pos_bal, num_days, cc_amt

def generate_working_sheet(
    accounts_data: list,
    wcdl_data: list,
    computed: dict,
    job_id: str,
    period: str,
    daily_loan_util: dict = None,  # Gap 4: {"WCDL": {date_str: float}, "BC": {...}, "PQL": {...}}
) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    used_names = {}
    for idx, account in enumerate(accounts_data):
        acct_no = str(account.get("account_number", "UNKNOWN")).strip()
        full_bank_name = account.get("bank_name", "UNKNOWN").upper()
        bank_short = full_bank_name.split()[0]
        acct_type = account.get("account_type", "CC").upper()
        
        # Determine a clean suffix from the account number
        acct_suffix = acct_no[-4:] if len(acct_no) >= 4 and acct_no != "UNKNOWN" else f"AC{idx+1}"
        
        # Consistent naming: HDFC-521 CC AC
        base_name = f"{bank_short}-{acct_suffix} {acct_type} AC"
        
        # Handle Excel's 31 character limit and uniqueness
        tab_name = base_name[:31]
        if tab_name in used_names:
            used_names[tab_name] += 1
            tab_name = f"{tab_name[:28]}_{used_names[tab_name]}"
        else:
            used_names[tab_name] = 1
            
        ws = wb.create_sheet(tab_name)
        
        # Row 1: Header - HDFC BANK - 521 CC Account + From 01-02-2026 to 28-02-2026
        period_from = account.get("period_from", "N/A")
        period_to = account.get("period_to", "N/A")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
        limit_val = account.get("cc_limit") or 0.0
        header_cell = ws.cell(row=1, column=1)
        header_cell.value = f"{full_bank_name} - {acct_suffix} {acct_type} Account | Limit: ₹{limit_val:,.2f} | From {period_from} to {period_to}"
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal="center")

        # Use dynamic names from scout where available, otherwise defaults
        h_narration = (account.get("narration_col_name") or "").strip() or "—"
        h_ref = (account.get("ref_col_name") or "").strip() or "—"
        h_wd = (account.get("withdrawal_col_name") or "").strip() or "—"
        h_dep = (account.get("deposit_col_name") or "").strip() or "—"
        h_bal = (account.get("balance_col_name") or "").strip() or "—"
        use_excel_balance_formulas = bool(account.get("use_excel_balance_formulas"))

        # Row 2: audit line — PDF column headers as scouted (not bank-name logic)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=13)
        pdf_audit = ws.cell(row=2, column=1)
        pdf_audit.value = (
            f"PDF columns: Date | {h_narration} | {h_ref} | {h_wd} | {h_dep} | {h_bal}"
            " — Row 2 = wording from the PDF (scout); Row 3 = fixed internal fields."
        )
        pdf_audit.font = Font(italic=True, size=9)
        pdf_audit.alignment = Alignment(horizontal="center", wrap_text=True)

        # Row 3: stable semantics for every bank (Debit/Credit from PDF map here as Dr/Cr).
        headers = [
            "Date",
            "Narration",
            "Chq/Ref No.",
            "Withdrawal (Dr)",
            "Deposit (Cr)",
            "Positive Bal",
            "No. of Days",
            "CC",
            "WCDL",
            "Buyers Credit",
            "Pre-Qualified Loan",
            "Total Utilisation",
            "Daily Interest",
        ]
        
        header_fill = PatternFill(fill_type="solid", fgColor="0A0A0F")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, title in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Styling: green only for calendar filler days (no PDF transactions that day).
        gap_fill = PatternFill(fill_type="solid", fgColor="E8F5E9")  # Lite Green
        
        # Transactions expansion
        txns = account.get("transactions", [])
        start_date = account.get("period_from")
        end_date = account.get("period_to")
        open_bal = account.get("opening_balance", 0)
        balance_style = account.get("balance_style") or "signed"

        full_month_txns = _expand_to_full_month(
            txns, start_date, end_date, open_bal, balance_style=balance_style
        )
        account["full_month_transactions"] = full_month_txns
        eod_row_for_date = _last_row_index_by_date(full_month_txns)

        roi = account.get("cc_roi_percent") or 7.60
        
        for i, txn in enumerate(full_month_txns):
            row_idx = i + 4
            
            if txn.get("is_gap"):
                for col_idx in range(1, 14):
                    ws.cell(row=row_idx, column=col_idx).fill = gap_fill

            # Date, Narration, Ref
            ws.cell(row=row_idx, column=1).value = txn.get("date")
            ws.cell(row=row_idx, column=2).value = txn.get("narration")
            ws.cell(row=row_idx, column=3).value = txn.get("ref_number")
            
            # Withdrawal (Column D) — prefer statement_transform withdrawal_dr
            wd_val = txn.get("withdrawal_dr")
            if wd_val is None:
                wd_val = txn.get("withdrawal") or 0.0
            wd_cell = ws.cell(row=row_idx, column=4)
            if txn.get("is_opening"):
                wd_cell.value = ""
            else:
                wd_cell.value = wd_val
            wd_cell.number_format = '#,##0.00'
            
            # Deposit (Column E) — prefer statement_transform deposit_cr
            dep_val = txn.get("deposit_cr")
            if dep_val is None:
                dep_val = txn.get("deposit") or 0.0
            dep_cell = ws.cell(row=row_idx, column=5)
            if txn.get("is_opening"):
                dep_cell.value = ""
            else:
                dep_cell.value = dep_val
            dep_cell.number_format = '#,##0.00'
            
            # Internal Truth (Column Z) - Keep it for reference
            # For the opening row, it's just open_bal. For others, it matches the txn cb.
            bal_cell = ws.cell(row=row_idx, column=26)
            cb = txn.get("closing_balance")
            if cb is None:
                cb = 0.0
            bal_cell.value = cb
            
            # F/G/H from blueprint balance_style + optional balance_credit_side (marker_inline Cr/OD).
            pos_bal_cell = ws.cell(row=row_idx, column=6)
            num_days_cell = ws.cell(row=row_idx, column=7)
            cc_cell = ws.cell(row=row_idx, column=8)

            side = txn.get("balance_credit_side")
            date_key = (txn.get("date") or "")[:10]
            is_end_of_day = bool(date_key) and eod_row_for_date.get(date_key) == i

            pos_bal, num_days, cc_amt = _positive_bal_num_days_cc(cb, balance_style, side)
            
            # F/G/H: populate for every row (user preference: process ALL rows)
            if use_excel_balance_formulas and side is None:
                pos_bal_cell.value = f'=IF(Z{row_idx}>0,Z{row_idx},0)'
                num_days_cell.value = f'=IF(F{row_idx}>0,1,0)'
                cc_cell.value = f'=IF(Z{row_idx}<0,ABS(Z{row_idx}),0)'
            else:
                pos_bal_cell.value = pos_bal
                num_days_cell.value = num_days
                cc_cell.value = cc_amt

            pos_bal_cell.number_format = '#,##0.00'
            cc_cell.number_format = '#,##0.00'
                
            # I: WCDL (from daily_loan_util if available, else 0)
            wcdl_val = 0.0
            if daily_loan_util:
                _date_key = txn.get("date", "")[:10]
                wcdl_val = float(daily_loan_util.get("WCDL", {}).get(_date_key, 0.0))
            ws.cell(row=row_idx, column=9).value = wcdl_val
            ws.cell(row=row_idx, column=9).number_format = '#,##0.00'

            # J: Buyers Credit (from daily_loan_util if available)
            bc_val = 0.0
            if daily_loan_util:
                _date_key = txn.get("date", "")[:10]
                bc_val = float(daily_loan_util.get("BC", {}).get(_date_key, 0.0))
            ws.cell(row=row_idx, column=10).value = bc_val
            ws.cell(row=row_idx, column=10).number_format = '#,##0.00'

            # K: Pre-Qualified Loan (from daily_loan_util if available)
            pql_val = 0.0
            if daily_loan_util:
                _date_key = txn.get("date", "")[:10]
                pql_val = float(daily_loan_util.get("PQL", {}).get(_date_key, 0.0))
            ws.cell(row=row_idx, column=11).value = pql_val
            ws.cell(row=row_idx, column=11).number_format = '#,##0.00'
            
            # L: Total Utilisation: H + I + J + K
            total_util_cell = ws.cell(row=row_idx, column=12)
            total_util_cell.value = f"=H{row_idx}+I{row_idx}+J{row_idx}+K{row_idx}"
            total_util_cell.number_format = '#,##0.00'
            
            # M: Daily Interest: (L * ROI / 365)
            daily_int_cell = ws.cell(row=row_idx, column=13)
            daily_int_cell.value = f"=(L{row_idx}*{roi}/100)/365"
            daily_int_cell.number_format = '#,##0.00'

        # Summary row
        last_row = len(full_month_txns) + 4
        ws.cell(row=last_row, column=1).value = "TOTAL"
        ws.cell(row=last_row, column=1).font = Font(bold=True)
        
        # Sum of CC, WCDL, BC, PQ, Total Util, Daily Int (H, I, J, K, L, M)
        for col_letter in ['H', 'I', 'J', 'K', 'L', 'M']:
            col_idx = ord(col_letter) - ord('A') + 1
            sum_cell = ws.cell(row=last_row, column=col_idx)
            sum_cell.value = f"=SUM({col_letter}4:{col_letter}{last_row-1})"
            sum_cell.font = Font(bold=True)
            sum_cell.number_format = '#,##0.00'
            
        # Avg Utilisation Row
        avg_row = last_row + 1
        ws.cell(row=avg_row, column=1).value = "AVG. UTILISATION"
        ws.cell(row=avg_row, column=1).font = Font(bold=True)
        avg_cell = ws.cell(row=avg_row, column=12) # Column L
        avg_cell.value = f"=AVERAGE(L4:L{last_row-1})"
        avg_cell.font = Font(bold=True)
        avg_cell.number_format = '#,##0.00'

        # Interest Reconciliation Block
        int_calc_row = avg_row + 3
        ws.cell(row=int_calc_row, column=1).value = "INTEREST RECONCILIATION"
        ws.cell(row=int_calc_row, column=1).font = Font(bold=True, size=11)
        
        ws.cell(row=int_calc_row+1, column=1).value = "Total CC Utilisation (Sum of Month)"
        ws.cell(row=int_calc_row+1, column=3).value = f"=SUM(H4:H{last_row-1})"
        
        ws.cell(row=int_calc_row+2, column=1).value = "Applicable ROI %"
        ws.cell(row=int_calc_row+2, column=3).value = f"={roi}/100"
        ws.cell(row=int_calc_row+2, column=3).number_format = '0.00%'

        ws.cell(row=int_calc_row+3, column=1).value = "Calculated Interest (SUM * ROI / 365)"
        ws.cell(row=int_calc_row+3, column=3).value = f"=(C{int_calc_row+1}*C{int_calc_row+2})/365"
        ws.cell(row=int_calc_row+3, column=3).font = Font(bold=True)
        ws.cell(row=int_calc_row+3, column=3).number_format = '#,##0.00'

        ws.cell(row=int_calc_row+3, column=4).value = "← Final CC Interest for the month"

        # Column Formatting
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['Z'].hidden = True # Internal balance
        for col in range(4, 15):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[col_letter].width = 16

    # Save file — use /tmp on Vercel (read-only filesystem), local path otherwise
    _is_vercel = os.getenv("VERCEL") or os.getenv("AWS_LAMBDA_FUNCTION_NAME")
    storage_dir = "/tmp/fincore/working_sheets" if _is_vercel else "./storage/working_sheets"
    os.makedirs(storage_dir, exist_ok=True)
    date_prefix = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{date_prefix}_workingsheet.xlsx"
    filepath = os.path.join(storage_dir, filename)
    wb.save(filepath)
    
    return filepath

def _parse_txn_date_key(t):
    d_str = (t.get("date") or "")[:10]
    if not d_str:
        return None
    try:
        if "-" in d_str and len(d_str) >= 10:
            return datetime.strptime(d_str[:10], "%Y-%m-%d").date()
        return datetime.strptime(d_str, "%d/%m/%Y").date()
    except Exception:
        return None


def _float_or_none(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _infer_balance_credit_side(cb: float, balance_style: str) -> Optional[bool]:
    if balance_style not in ("marker_inline", "dr_cr_suffix"):
        return None
    if cb > 0:
        return True
    if cb < 0:
        return False
    return None


def _expand_to_full_month(transactions, start_str, end_str, open_bal, balance_style: str = "signed"):
    """
    One row per calendar day in [period_from, period_to].

    - PDF transactions for a day are emitted in order with real=True.
    - Days with no PDF rows get a single carry-forward row (green in the sheet via is_gap).
    - No separate synthetic opening row on period_from (avoids duplicating that date).
    - Running balance: prefer each txn's closing_balance when present; else Dr/Cr roll-forward.
    """
    if not start_str or not end_str:
        return transactions

    try:
        start_date = datetime.strptime(start_str[:10], "%Y-%m-%d").date()
        end_date = datetime.strptime(end_str[:10], "%Y-%m-%d").date()
    except Exception:
        return transactions

    from collections import defaultdict

    txns_by_date = defaultdict(list)
    for t in transactions:
        dt = _parse_txn_date_key(t)
        if dt is not None:
            txns_by_date[dt].append(t)

    expanded = []
    try:
        current_bal = float(open_bal) if open_bal is not None else 0.0
    except (TypeError, ValueError):
        current_bal = 0.0

    curr = start_date
    while curr <= end_date:
        day_txns = txns_by_date.get(curr, [])
        if day_txns:
            for t in day_txns:
                t["real"] = True
                t["is_gap"] = False
                if "is_opening" in t:
                    del t["is_opening"]
                wd = float(t.get("withdrawal") or 0)
                dp = float(t.get("deposit") or 0)
                cb = _float_or_none(t.get("closing_balance"))
                if cb is not None:
                    current_bal = cb
                else:
                    current_bal = current_bal - wd + dp
                t["closing_balance"] = current_bal
                if t.get("balance_credit_side") is None:
                    inf = _infer_balance_credit_side(current_bal, balance_style)
                    if inf is not None:
                        t["balance_credit_side"] = inf
                expanded.append(t)
        else:
            expanded.append(
                {
                    "date": curr.strftime("%Y-%m-%d"),
                    "narration": "NO TRANSACTION FOR THE DAY",
                    "ref_number": "-",
                    "withdrawal": 0.0,
                    "deposit": 0.0,
                    "closing_balance": current_bal,
                    "balance_credit_side": _infer_balance_credit_side(
                        current_bal, balance_style
                    ),
                    "category": "INTERNAL",
                    "real": False,
                    "is_gap": True,
                }
            )
        curr += timedelta(days=1)

    return expanded

def _populate_shared_tab(ws, data, title):
    """ Simple aggregation for Charges, Interest, etc. """
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(bold=True, size=12)
    
    headers = ["Date", "Account", "Particulars", "Amount", "Dr/Cr"]
    for col, h in enumerate(headers, 1):
        ws.cell(3, col).value = h
        ws.cell(3, col).font = Font(bold=True)
        
    for row_idx, item in enumerate(data, 4):
        amt = item.get("withdrawal") or item.get("deposit") or 0
        flag = "Dr" if item.get("withdrawal") else "Cr"
        
        ws.cell(row_idx, 1).value = item.get("date")
        ws.cell(row_idx, 2).value = item.get("acct")
        ws.cell(row_idx, 3).value = item.get("narration")
        ws.cell(row_idx, 4).value = amt
        ws.cell(row_idx, 5).value = flag

def _populate_wcdl_tracker(ws, wcdl_data, engine):
    headers = [
        "Loan Number", "Drawdown Date",
        "Maturity Date", "ROI %",
        "Principal", "Tenure Days",
        "Prepayment Date", "Actual Tenure",
        "Computed Interest", "Bank Charged",
        "Difference", "Status"
    ]
    
    header_fill = PatternFill(fill_type="solid", fgColor="BCF0F0")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
    
    for row_idx, loan in enumerate(wcdl_data, 2):
        computed = engine.compute_wcdl_interest(
            loan.get("principal", 0),
            loan.get("roi_percent", 0),
            loan.get("tenure_days", 0)
        )
        
        bank_charged = loan.get("bank_charged_interest", computed)
        difference = computed - bank_charged
        status = "✓ MATCH" if abs(difference) <= 1 else "⚠ FLAG"
        
        row_data = [
            loan.get("loan_number", "WCDL-TXN"),
            loan.get("date", ""),
            loan.get("maturity_date", ""),
            loan.get("roi_percent", 0),
            loan.get("principal", 0),
            loan.get("tenure_days", 0),
            loan.get("prepayment_date", ""),
            loan.get("actual_tenure", loan.get("tenure_days", 0)),
            computed,
            bank_charged,
            round(difference, 2),
            status
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col).value = value

def _populate_summary_dashboard(ws, computed, period, account_count):
    # Snapshot of the analytics
    ws.cell(1, 1).value = f"Financial Summary — {period}"
    ws.cell(1, 1).font = Font(size=14, bold=True)
    
    kpis = [
        ("Accounts Processed", account_count),
        ("Average Utilization", computed.get("average_utilisation", 0)),
        ("Total Interest (CC + WCDL)", computed.get("total_interest", 0)),
        ("Finance Cost (%)", f"{computed.get('finance_cost_pct', 0):.2f}%"),
        ("ROI Status", computed.get("roi_status", "Checking...")),
    ]
    
    for idx, (label, value) in enumerate(kpis, 3):
        ws.cell(idx, 1).value = label
        ws.cell(idx, 2).value = value
        ws.cell(idx, 1).font = Font(bold=True)
        
    # Styling
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
