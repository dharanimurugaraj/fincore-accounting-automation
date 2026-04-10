"""
fx_sheet.py — FX Account Working Sheet Generator

Gap 7 Fix: Handles FX (USD / EUR / GBP) accounts.

For each FX account (account_type == "FX"):
  - date
  - FC balance (foreign currency)
  - No. of Days
  - Exchange rate (from statement or config)
  - INR equivalent (FC balance × exchange rate)

Bottom rows:
  - Total days with balance
  - Avg FC balance
  - Avg INR equivalent

One Excel tab per currency (named "FX-USD", "FX-EUR", etc.)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from typing import Any, Dict, List, Optional


def generate_fx_sheet(
    fx_accounts_data: List[Dict[str, Any]],
    job_id: str,
    period: str,
) -> str:
    """
    Generate a multi-tab Excel workbook for FX accounts.

    Args:
        fx_accounts_data: list of extracted account dicts where account_type == "FX".
                          Each dict must include 'currency', 'transactions', etc.
        job_id:           pipeline run ID (for filename uniqueness)
        period:           human-readable period label e.g. "Feb-2026"

    Returns:
        Absolute path to the saved .xlsx file.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    used_tabs: Dict[str, int] = {}

    for account in fx_accounts_data:
        currency    = (account.get("currency") or "FC").upper()
        bank_name   = (account.get("bank_name") or "UNKNOWN").upper()
        acct_no     = str(account.get("account_number") or "").strip()
        acct_suffix = acct_no[-4:] if len(acct_no) >= 4 else acct_no or "FX"
        period_from = account.get("period_from", "N/A")
        period_to   = account.get("period_to", "N/A")

        # Tab name e.g. "FX-USD (HDFC)", max 31 chars
        base_tab = f"FX-{currency} ({bank_name.split()[0]})"[:28]
        tab_name = base_tab
        if tab_name in used_tabs:
            used_tabs[tab_name] += 1
            tab_name = f"{base_tab}_{used_tabs[tab_name]}"
        else:
            used_tabs[tab_name] = 1

        ws = wb.create_sheet(tab_name)
        _build_fx_tab(ws, account, currency, bank_name, acct_suffix, period_from, period_to, period)

    # Save
    is_vercel   = os.getenv("VERCEL") or os.getenv("AWS_LAMBDA_FUNCTION_NAME")
    storage_dir = "/tmp/fincore/fx_sheets" if is_vercel else "./storage/fx_sheets"
    os.makedirs(storage_dir, exist_ok=True)
    date_prefix = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename    = f"{date_prefix}_fx_{job_id[-6:]}.xlsx"
    filepath    = os.path.join(storage_dir, filename)
    wb.save(filepath)
    return filepath


# ── Internal builder ──────────────────────────────────────────────────────────

def _build_fx_tab(
    ws,
    account: Dict[str, Any],
    currency: str,
    bank_name: str,
    acct_suffix: str,
    period_from: str,
    period_to: str,
    period: str,
) -> None:
    """Populate one FX tab with daily FC balance, exchange rate, and INR equivalent."""

    header_fill = PatternFill(fill_type="solid", fgColor="0A3055")   # Dark navy
    header_font = Font(bold=True, color="FFFFFF")
    alt_fill    = PatternFill(fill_type="solid", fgColor="EEF4FA")
    total_fill  = PatternFill(fill_type="solid", fgColor="D0E8FF")
    thin        = Side(style="thin", color="B0C4DE")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Title ─────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    title_cell        = ws.cell(row=1, column=1)
    title_cell.value  = (
        f"{bank_name} | {currency} FX Account ({acct_suffix}) "
        f"| {period_from} to {period_to}"
    )
    title_cell.font       = Font(bold=True, size=12, color="0A3055")
    title_cell.alignment  = Alignment(horizontal="center")

    # ── Row 3: Column headers ──────────────────────────────────────────────
    col_headers = [
        "Date",
        f"FC Balance ({currency})",
        "No. of Days",
        "Exchange Rate (₹/FC)",
        "INR Equivalent (₹)",
        "Remarks",
    ]
    for col, h in enumerate(col_headers, 1):
        cell           = ws.cell(row=3, column=col)
        cell.value     = h
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border    = border

    # ── Expand transactions to full calendar period ────────────────────────
    from .working_sheet import _expand_to_full_month

    txns = account.get("transactions", [])
    open_bal = float(account.get("opening_balance") or 0)
    full_txns = _expand_to_full_month(
        txns,
        period_from,
        period_to,
        open_bal,
        balance_style=account.get("balance_style") or "signed",
    )

    # ── Data rows ─────────────────────────────────────────────────────────
    data_start = 4
    for i, txn in enumerate(full_txns):
        row_idx  = data_start + i
        is_real  = txn.get("real", True)
        fill     = None if is_real else alt_fill

        date_val = txn.get("date", "")
        # FC balance = closing_balance from statement (already in FC units)
        fc_bal   = float(txn.get("closing_balance") or 0)
        # Exchange rate: pulled from txn if scouted, else defaults to 0 (manual entry)
        fx_rate  = float(txn.get("exchange_rate") or txn.get("fx_rate") or 0)
        # No. of Days: 1 if there is any balance, else 0
        no_days  = 1 if fc_bal != 0 else 0
        narration = txn.get("narration", "")

        row_cells = [
            (1, date_val,    None),
            (2, fc_bal,      "#,##0.0000"),
            (3, no_days,     "0"),
            (4, fx_rate,     "#,##0.4f"),
            (5, None,        "#,##0.00"),   # Formula: =B{r}*D{r}
            (6, narration,   None),
        ]

        for col, val, fmt in row_cells:
            cell = ws.cell(row=row_idx, column=col)
            if col == 5:
                # INR Equivalent formula
                b_col = get_column_letter(2)
                d_col = get_column_letter(4)
                cell.value = f"={b_col}{row_idx}*{d_col}{row_idx}"
            else:
                cell.value = val
            if fmt:
                cell.number_format = fmt
            if fill:
                cell.fill = fill
            cell.border = border

    # ── Summary / Total rows ──────────────────────────────────────────────
    last_data_row = data_start + len(full_txns) - 1
    total_row     = last_data_row + 2
    avg_row       = total_row + 1

    # TOTAL row
    ws.cell(row=total_row, column=1).value = "TOTAL / SUM"
    ws.cell(row=total_row, column=1).font  = Font(bold=True)
    ws.cell(row=total_row, column=1).fill  = total_fill

    for col_letter, col_idx in [("B", 2), ("C", 3), ("E", 5)]:
        cell = ws.cell(row=total_row, column=col_idx)
        cell.value        = f"=SUM({col_letter}{data_start}:{col_letter}{last_data_row})"
        cell.font         = Font(bold=True)
        cell.fill         = total_fill
        cell.number_format = "#,##0.00"
        cell.border        = border

    # AVG row
    ws.cell(row=avg_row, column=1).value = "AVG UTILISATION"
    ws.cell(row=avg_row, column=1).font  = Font(bold=True)
    ws.cell(row=avg_row, column=1).fill  = total_fill

    for col_letter, col_idx in [("B", 2), ("E", 5)]:
        cell = ws.cell(row=avg_row, column=col_idx)
        cell.value        = f"=AVERAGE({col_letter}{data_start}:{col_letter}{last_data_row})"
        cell.font         = Font(bold=True)
        cell.fill         = total_fill
        cell.number_format = "#,##0.00"
        cell.border        = border

    # Avg in FC label
    ws.cell(row=avg_row, column=3).value = f"Avg FC ({currency})"
    ws.cell(row=avg_row, column=4).value = "→ Avg INR Equiv"

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 35


def build_daily_loan_util_dict(
    loans,
    date_range,
) -> Dict[str, Dict[str, float]]:
    """
    Utility: Build the daily_loan_util dict consumed by generate_working_sheet().

    Args:
        loans:      list of LoanTracker instances
        date_range: list of date objects (calendar period)

    Returns:
        {
            "WCDL": {"2026-02-01": 300000000.0, ...},
            "BC":   {"2026-02-01": 50000000.0,  ...},
            "PQL":  {"2026-02-01": 0.0,          ...},
        }
    """
    from .loan_tracker import get_wcdl_utilisation, get_bc_utilisation, get_pql_utilisation

    wcdl_util = get_wcdl_utilisation(loans, date_range)
    bc_util   = get_bc_utilisation(loans, date_range)
    pql_util  = get_pql_utilisation(loans, date_range)

    def _to_str_keys(d):
        return {str(k): v for k, v in d.items()}

    return {
        "WCDL": _to_str_keys(wcdl_util),
        "BC":   _to_str_keys(bc_util),
        "PQL":  _to_str_keys(pql_util),
    }
