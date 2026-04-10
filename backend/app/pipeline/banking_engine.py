import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import logging

# Configure Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger("BankingEngine")

class BankConfigMappingError(Exception):
    """Raised when a required column is missing or renamed in the bank statement."""
    pass

# --- STEP 1: Bank Config Registry ---
BANK_CONFIG_REGISTRY = {
    "HDFC-521": {
        "bank_name": "HDFC-521",
        "account_number": "521",
        "account_type": "CC",
        "currency": "INR",
        "col_date": "Date",
        "date_format": "%d/%m/%Y",
        "col_closing_balance": "Balance",
        "col_withdrawal": "Withdrawal (Dr)",
        "col_deposit": "Deposit (Cr)",
        "balance_sign_convention": "flagged", # Separate Dr/Cr column
        "col_dr_cr": "Dr/Cr Flag", # Placeholder - will be matched exactly from file
        "dr_value": "Dr",
        "cc_roi": 0.0725,
        "cc_sanctioned_limit": 225000000, # 22.5 Cr
        "total_wc_limit": 500000000,
    },
    "HDFC-512": {
        "bank_name": "HDFC-512",
        "account_number": "512",
        "account_type": "CA",
        "currency": "INR",
        "col_date": "Date",
        "date_format": "%d/%m/%Y",
        "col_closing_balance": "Balance",
        "col_withdrawal": "Debit",
        "col_deposit": "Credit",
        "balance_sign_convention": "signed", # negative = Dr
        "col_dr_cr": None,
        "dr_value": None,
        "cc_roi": 0.0,
        "cc_sanctioned_limit": 0,
        "total_wc_limit": 500000000,
    },
    "HDFC-552": {
        "bank_name": "HDFC-552",
        "account_number": "552",
        "account_type": "CA",
        "currency": "INR",
        "col_date": "Date",
        "date_format": "%d/%m/%Y",
        "col_closing_balance": "Balance",
        "col_withdrawal": "Debit",
        "col_deposit": "Credit",
        "balance_sign_convention": "signed",
        "col_dr_cr": None,
        "dr_value": None,
        "cc_roi": 0.0,
        "cc_sanctioned_limit": 0,
        "total_wc_limit": 500000000,
    },
    "UBI": {
        "bank_name": "UBI",
        "account_number": "UBI-001",
        "account_type": "CC",
        "currency": "INR",
        "col_date": "Txn Date",
        "date_format": "%d-%b-%y",
        "col_closing_balance": "Closing Bal",
        "col_withdrawal": "Withdrawal",
        "col_deposit": "Deposit",
        "balance_sign_convention": "signed",
        "col_dr_cr": None,
        "dr_value": "Dr",
        "cc_roi": 0.0810,
        "cc_sanctioned_limit": 150000000,
        "total_wc_limit": 500000000,
    }
}

# --- STEP 3: Core Calculation Engine ---

class DailyRow:
    def __init__(self, date: datetime, closing_balance: float, withdrawal: float = 0.0, deposit: float = 0.0):
        self.date = date
        self.closing_balance = closing_balance
        self.withdrawal = withdrawal
        self.deposit = deposit
        self.cc_utilisation = abs(closing_balance) if closing_balance < 0 else 0.0
        self.positive_balance = closing_balance if closing_balance > 0 else 0.0
        self.no_of_days_positive = 1 if closing_balance > 0 else 0

def parse_statement(file_path: str, bank_config: Dict[str, Any]) -> List[DailyRow]:
    """
    Read file, normalise columns mapping, determine balance sign, and fill missing dates.
    """
    ext = os.path.splitext(file_path)[1].lower()
    try:
        if ext == '.csv':
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
    except Exception as e:
        raise BankConfigMappingError(f"Failed to read file {file_path}: {str(e)}")

    col_date = bank_config["col_date"]
    col_bal = bank_config["col_closing_balance"]
    col_wd = bank_config.get("col_withdrawal")
    col_dep = bank_config.get("col_deposit")
    
    if col_date not in df.columns:
        raise BankConfigMappingError(f"col_date '{col_date}' not found in statement")
    if col_bal not in df.columns:
        raise BankConfigMappingError(f"col_closing_balance '{col_bal}' not found in statement")

    # Date Normalisation
    if pd.api.types.is_datetime64_any_dtype(df[col_date]):
        df['parsed_date'] = df[col_date]
    else:
        df['parsed_date'] = pd.to_datetime(df[col_date], format=bank_config["date_format"], dayfirst=True, errors='coerce')
    
    df = df.dropna(subset=['parsed_date'])

    # Balance Normalisation
    if bank_config["balance_sign_convention"] == "flagged":
        col_dr_cr = bank_config["col_dr_cr"]
        if col_dr_cr and col_dr_cr in df.columns:
            dr_val = bank_config["dr_value"]
            df['norm_balance'] = df.apply(
                lambda row: -abs(float(row[col_bal])) if str(row[col_dr_cr]).strip() == dr_val else abs(float(row[col_bal])),
                axis=1
            )
        else:
            # Fallback if col_dr_cr is missing but flagged was expected
            logger.warning(f"Flagged convention specified but col_dr_cr '{col_dr_cr}' not found. Using raw values.")
            df['norm_balance'] = df[col_bal].astype(float)
    else:
        # signed convention
        df['norm_balance'] = df[col_bal].astype(float)

    # Aggregate columns
    agg_dict = {'norm_balance': 'last'}
    if col_wd and col_wd in df.columns: agg_dict[col_wd] = 'sum'
    if col_dep and col_dep in df.columns: agg_dict[col_dep] = 'sum'
    
    daily_agg = df.groupby('parsed_date').agg(agg_dict).reset_index()
    
    # Gap Filling (Carry Forward)
    start_date = daily_agg['parsed_date'].min()
    end_date = daily_agg['parsed_date'].max()
    
    # Generate full date range
    all_dates = pd.date_range(start=start_date, end=end_date, freq='D')
    daily_full = daily_agg.set_index('parsed_date').reindex(all_dates).reset_index()
    daily_full.columns = ['date', 'balance'] + [c for c in agg_dict.keys() if c != 'norm_balance']
    
    # Carry forward logic for balance
    daily_full['balance'] = daily_full['balance'].ffill().fillna(0.0)
    # Withdrawals/Deposits should be 0 for missing dates
    for c in daily_full.columns:
        if c not in ['date', 'balance']:
            daily_full[c] = daily_full[c].fillna(0.0)

    rows = []
    for _, r in daily_full.iterrows():
        wd_val = r.get(col_wd, 0.0) if col_wd else 0.0
        dep_val = r.get(col_dep, 0.0) if col_dep else 0.0
        rows.append(DailyRow(r['date'], r['balance'], wd_val, dep_val))
    
    return rows

def get_cc_utilisation(daily_rows: List[DailyRow]) -> Dict[str, Any]:
    total_days = len(daily_rows)
    if total_days == 0: return {"avg_cc": 0, "sum_cc": 0}
    sum_cc = sum(r.cc_utilisation for r in daily_rows)
    return {
        "avg_cc": sum_cc / total_days,
        "sum_cc": sum_cc,
        "total_days": total_days
    }

def get_loan_utilisation(loan_tracker: List[Dict[str, Any]], date_range: List[datetime], loan_type: str) -> Dict[datetime, float]:
    """
    Returns a mapping of date -> total outstanding principal for specified loan_type (WCDL / BC / PQL)
    """
    results = {}
    for dt in date_range:
        day_total = 0.0
        for loan in loan_tracker:
            if loan.get("loan_type") != loan_type: continue
            
            drawdown = pd.to_datetime(loan["drawdown_date"])
            maturity = pd.to_datetime(loan["maturity_date"])
            end_date = maturity
            if loan.get("prepayment_date"):
                prepay = pd.to_datetime(loan["prepayment_date"])
                end_date = min(maturity, prepay)
            
            if drawdown <= dt <= end_date:
                day_total += float(loan.get("principal_inr", 0))
        results[dt] = day_total
    return results

def calculate_interest(facility_type: str, daily_balances: List[float], roi: float) -> Dict[str, Any]:
    """
    CC interest: SUM(daily_cc_balances) * roi / 365
    """
    total_days = len(daily_balances)
    if total_days == 0: return {"interest": 0, "eff_roi": 0}
    
    sum_bal = sum(daily_balances)
    interest = (sum_bal * roi) / 365
    avg_util = sum_bal / total_days
    
    eff_roi = (interest / avg_util * 365 / total_days) if avg_util > 0 else 0
    
    return {
        "calculated_interest": round(interest, 2),
        "effective_roi_annualised": round(eff_roi, 6),
        "avg_utilisation": round(avg_util, 2)
    }

# --- STEP 4: Excel Output Generation ---

def build_working_sheet(wb: openpyxl.Workbook, bank_id: str, daily_rows: List[DailyRow], loan_tracker: List[Dict[str, Any]]):
    config = BANK_CONFIG_REGISTRY[bank_id]
    ws = wb.create_sheet(f"{config['bank_name']} - {config['account_number']}")
    
    # Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # DYNAMIC HEADERS
    wd_header = config.get("col_withdrawal", "Withdrawal")
    dep_header = config.get("col_deposit", "Deposit")
    
    headers = [
        "Date", wd_header, dep_header, "Closing Balance (INR)", "Positive Bal", 
        "No of Days", "CC Utilisation", "WCDL", "Buyer's Credit", "PQL", "Total Utilisation"
    ]
    
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    date_range = [r.date for r in daily_rows]
    wcdl_vals = get_loan_utilisation(loan_tracker, date_range, "WCDL")
    bc_vals = get_loan_utilisation(loan_tracker, date_range, "BC")
    pql_vals = get_loan_utilisation(loan_tracker, date_range, "PQL")

    for i, row in enumerate(daily_rows, 2):
        dt = row.date
        w = wcdl_vals.get(dt, 0)
        b = bc_vals.get(dt, 0)
        p = pql_vals.get(dt, 0)
        total_util = row.cc_utilisation + w + b + p
        
        data = [
            dt.strftime("%Y-%m-%d"), row.withdrawal, row.deposit, row.closing_balance, 
            row.positive_balance, row.no_of_days_positive, row.cc_utilisation, w, b, p, total_util
        ]
        
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if col > 1: cell.number_format = '#,##0.00'

    # Summary Rows
    last_data_row = len(daily_rows) + 1
    total_row = last_data_row + 1
    avg_row = last_data_row + 2
    
    ws.cell(total_row, 1, "TOTAL").font = Font(bold=True)
    ws.cell(avg_row, 1, "AVG UTILISATION").font = Font(bold=True)
    
    for col in range(2, 12):
        col_letter = get_column_letter(col)
        # Total
        ws.cell(total_row, col, f"=SUM({col_letter}2:{col_letter}{last_data_row})").font = Font(bold=True)
        ws.cell(total_row, col).number_format = '#,##0.00'
        # Avg
        ws.cell(avg_row, col, f"=AVERAGE({col_letter}2:{col_letter}{last_data_row})").font = Font(bold=True)
        ws.cell(avg_row, col).number_format = '#,##0.00'

    # Interest Section
    ws.cell(avg_row + 4, 1, "INTEREST CALCULATIONS").font = Font(bold=True, size=12)
    roi = config.get("cc_roi", 0.0725)
    
    sum_cc = sum(r.cc_utilisation for r in daily_rows)
    calc_int = (sum_cc * roi) / 365
    
    ws.cell(avg_row + 5, 1, "Annual ROI")
    ws.cell(avg_row + 5, 2, roi).number_format = '0.00%'
    ws.cell(avg_row + 6, 1, "Calculated Monthly CC Interest")
    ws.cell(avg_row + 6, 2, calc_int).number_format = '#,##0.00'

    # WCDL Verification Table
    ws.cell(avg_row + 8, 1, "WCDL VERIFICATION").font = Font(bold=True)
    w_headers = ["Loan Ref", "Drawdown", "Principal", "ROI", "Our Calc", "Bank Charge", "Diff"]
    for col, h in enumerate(w_headers, 1):
        ws.cell(avg_row + 9, col, h).font = Font(bold=True)
    
    # ... Populate WCDL loans specific to this bank ...

def build_consolidated_detail(wb: openpyxl.Workbook, all_data: Dict[str, List[DailyRow]], loan_tracker: List[Dict[str, Any]]):
    ws = wb.create_sheet("Daily Detail Sheet")
    
    # Get overall date range
    all_dates_set = set()
    for rows in all_data.values():
        for r in rows: all_dates_set.add(r.date)
    sorted_dates = sorted(list(all_dates_set))
    
    headers = ["Date"]
    # Bank CC columns
    for bank_id in all_data.keys():
        headers.append(f"{bank_id} CC")
    
    headers += ["Total CC Util", "Total WCDL", "Total BC", "Total PQL", "Grand Total Util", "Daily Interest Accrual"]
    
    for col, text in enumerate(headers, 1):
        ws.cell(1, col, text).font = Font(bold=True)
        ws.cell(1, col).fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    for i, dt in enumerate(sorted_dates, 2):
        ws.cell(i, 1, dt.strftime("%Y-%m-%d"))
        col_ptr = 2
        total_cc = 0
        for bank_id, rows in all_data.items():
            # Find row for this date
            day_cc = 0
            for r in rows:
                if r.date == dt:
                    day_cc = r.cc_utilisation
                    break
            ws.cell(i, col_ptr, day_cc).number_format = '#,##0.00'
            total_cc += day_cc
            col_ptr += 1
        
        # Loans
        l_w = get_loan_utilisation(loan_tracker, [dt], "WCDL")[dt]
        l_b = get_loan_utilisation(loan_tracker, [dt], "BC")[dt]
        l_p = get_loan_utilisation(loan_tracker, [dt], "PQL")[dt]
        
        ws.cell(i, col_ptr, total_cc).number_format = '#,##0.00'
        ws.cell(i, col_ptr+1, l_w).number_format = '#,##0.00'
        ws.cell(i, col_ptr+2, l_b).number_format = '#,##0.00'
        ws.cell(i, col_ptr+3, l_p).number_format = '#,##0.00'
        
        grand_total = total_cc + l_w + l_b + l_p
        ws.cell(i, col_ptr+4, grand_total).number_format = '#,##0.00'
        
        # Sample Daily Accrual (using HDFC ROI 7.25% as blend placeholder)
        ws.cell(i, col_ptr+5, (grand_total * 0.0725 / 365)).number_format = '#,##0.00'

def build_wcdl_tracker_tab(wb: openpyxl.Workbook, loan_tracker: List[Dict[str, Any]]):
    ws = wb.create_sheet("WCDL Tracker")
    headers = ["Loan Ref", "Bank", "Drawdown", "Maturity", "Prepay Date", "Principal (INR)", "ROI", "Status"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h).font = Font(bold=True)
    
    for i, loan in enumerate(loan_tracker, 2):
        ws.cell(i, 1, loan.get("loan_ref"))
        ws.cell(i, 2, loan.get("bank_name"))
        ws.cell(i, 3, loan.get("drawdown_date"))
        ws.cell(i, 4, loan.get("maturity_date"))
        ws.cell(i, 5, loan.get("prepayment_date"))
        ws.cell(i, 6, loan.get("principal_inr")).number_format = '#,##0.00'
        ws.cell(i, 7, loan.get("loan_roi")).number_format = '0.00%'
        
        # Status logic
        active = True
        if loan.get("prepayment_date"): active = False
        ws.cell(i, 8, "Active" if active else "Closed")

def build_fx_sheet(wb: openpyxl.Workbook, bank_id: str, daily_rows: List[DailyRow]):
    config = BANK_CONFIG_REGISTRY[bank_id]
    ws = wb.create_sheet(f"FX-{config['currency']}-{config['account_number']}")
    headers = ["Date", "Balance (FC)", "Exchange Rate", "Balance (INR)", "No of Days"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h).font = Font(bold=True)
    
    for i, row in enumerate(daily_rows, 2):
        ws.cell(i, 1, row.date.strftime("%Y-%m-%d"))
        ws.cell(i, 2, row.closing_balance).number_format = '#,##0.00'
        # Placeholder Rate (Should ideally come from config or data)
        rate = 83.0 if config['currency'] == 'USD' else (90.0 if config['currency'] == 'EUR' else 1.0)
        ws.cell(i, 3, rate)
        ws.cell(i, 4, row.closing_balance * rate).number_format = '#,##0.00'
        ws.cell(i, 5, row.no_of_days_positive)

def build_summary_report(wb: openpyxl.Workbook, all_data: Dict[str, List[DailyRow]], loan_tracker: List[Dict[str, Any]], month_year: str):
    ws = wb.create_sheet("Banking Report", 0)
    
    # Section A - WC Utilisation
    ws.cell(1, 1, "Section A — WC Utilisation").font = Font(bold=True, size=12)
    headers = ["Facility", "Avg Utilisation (INR)", "Total Sanction (INR)", "% Utilisation"]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h).font = Font(bold=True)
    
    # Calculate totals
    total_avg_cc = sum(get_cc_utilisation(rows)["avg_cc"] for rows in all_data.values())
    
    # Get overall date range for loans
    all_dates = set()
    for rows in all_data.values():
        for r in rows: all_dates.add(r.date)
    sorted_dates = sorted(list(all_dates))
    
    wcdl_daily = get_loan_utilisation(loan_tracker, sorted_dates, "WCDL")
    avg_wcdl = sum(wcdl_daily.values()) / len(sorted_dates) if sorted_dates else 0
    
    # Total Sanction
    total_sanction = sum(config.get("total_wc_limit", 0) for config in BANK_CONFIG_REGISTRY.values())
    
    facilities = [
        ("Cash Credit (CC)", total_avg_cc),
        ("WCDL / BC / PQL", avg_wcdl),
        ("Total", total_avg_cc + avg_wcdl)
    ]
    
    for i, (f_name, val) in enumerate(facilities, 3):
        ws.cell(i, 1, f_name)
        ws.cell(i, 2, val).number_format = '#,##0.00'
        ws.cell(i, 3, total_sanction if f_name == "Total" else "")
        if f_name == "Total" and total_sanction > 0:
            ws.cell(i, 4, val / total_sanction).number_format = '0.00%'

    # Section B - Interest Cost
    ws.cell(8, 1, "Section A (Right) — Interest Cost").font = Font(bold=True)
    # ... more summary fields ...

def generate_banking_report_package(statements: List[Dict[str, str]], loan_tracker: List[Dict[str, Any]], month_year: str) -> str:
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
    
    all_parsed_rows = {}
    
    for st in statements:
        bid = st["bank_id"]
        path = st["path"]
        if bid not in BANK_CONFIG_REGISTRY:
            logger.error(f"Bank ID {bid} not in registry!")
            continue
            
        daily_rows = parse_statement(path, BANK_CONFIG_REGISTRY[bid])
        all_parsed_rows[bid] = daily_rows
        
        # Determine if it's FX or regular
        if BANK_CONFIG_REGISTRY[bid]["currency"] != "INR":
            build_fx_sheet(wb, bid, daily_rows)
        else:
            build_working_sheet(wb, bid, daily_rows, loan_tracker)

    build_wcdl_tracker_tab(wb, loan_tracker)
    build_consolidated_detail(wb, all_parsed_rows, loan_tracker)
    build_summary_report(wb, all_parsed_rows, loan_tracker, month_year)

    _is_vercel = os.getenv("VERCEL") or os.getenv("AWS_LAMBDA_FUNCTION_NAME")
    output_dir = "/tmp/fincore/reports" if _is_vercel else "storage/reports"
    os.makedirs(output_dir, exist_ok=True)
    filename = f"Banking_Performance_Report_{month_year.replace('/', '_')}.xlsx"
    file_path = os.path.join(output_dir, filename)
    wb.save(file_path)
    return file_path
