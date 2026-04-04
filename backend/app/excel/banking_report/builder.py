"""
Banking Report Generator — reads Working Sheet, produces a Banking Report.

All tabs and sections are DYNAMIC — they depend on what sheets exist in the
Working Sheet (which depends on what PDFs were uploaded).

TAB 1: "<Month>" — Management Snapshot (utilisation, finance cost, charges, ROI)
TAB 2: "Detail Sheet" — Daily Consolidated View across all accounts
TAB 3: "Forex" — Import Transaction Register (only if forex data exists)
"""

import datetime
import calendar
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, List, Any, Optional

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style constants ───────────────────────────────────────────────────────────

HDR_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HDR_FILL = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
SECTION_FONT = Font(name="Arial", bold=True, size=11, color="0D1B2A")
SECTION_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
LABEL_FONT = Font(name="Arial", size=9, color="333333")
VALUE_FONT = Font(name="Arial", bold=True, size=10)
ALERT_FONT = Font(name="Arial", bold=True, size=10, color="C00000")
DATA_FONT = Font(name="Arial", size=9)
NUM_FMT = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
ALT_FILL = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

CC_KEYWORDS = {"cc", "cash credit", "od", "overdraft"}
CURRENT_KEYWORDS = {"current", "savings", "ubi"}
SKIP_SHEETS = {"Interest", "Charges", "EMI", "WCDL Tracker", "Forex", "No Data"}


def generate_banking_report(
    working_sheet_path: str,
    forex_rates: dict,
    market_rates: dict,
    output_path: str,
) -> str:
    """
    Reads the Working Sheet and generates a dynamically-structured Banking Report.
    """
    ws_wb = load_workbook(working_sheet_path, data_only=True)
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    data = _extract_ws_data(ws_wb)

    _build_snapshot_tab(out_wb, data)
    _build_detail_sheet_tab(out_wb, data)

    if data.get("has_forex"):
        _build_forex_tab(out_wb, data, market_rates)

    out_wb.save(output_path)
    return output_path


# ── Tab 1: Management Snapshot ────────────────────────────────────────────────

def _build_snapshot_tab(wb: Workbook, data: dict):
    month_label = data.get("month_label", "Report")
    ws = wb.create_sheet(month_label[:31])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22

    row = 1

    # Section A: Utilisation
    row = _write_section_header(ws, row, "A. Utilisation Summary")
    for label, val in [
        ("Total CC Avg Utilisation", _fmt_num(data.get("avg_cc_util_inr", 0))),
        ("Total WCDL Avg Utilisation", _fmt_num(data.get("avg_wcdl_util_inr", 0))),
        ("Total Avg Utilisation", _fmt_num(data.get("avg_total_util_inr", 0))),
    ]:
        row = _write_kv_row(ws, row, label, val)
    row += 1

    # Finance Cost
    row = _write_section_header(ws, row, "A. Finance Cost")
    for label, val in [
        ("Total Interest Charged", _fmt_num(data.get("total_interest", 0))),
        ("Finance Cost % (Annualised)", _fmt_pct(data.get("finance_cost_pct", 0))),
    ]:
        row = _write_kv_row(ws, row, label, val)
    row += 1

    # Section B: Charges
    ch = data.get("charges", {})
    if ch.get("total", 0) > 0 or ch.get("penal", 0) > 0:
        row = _write_section_header(ws, row, "B. Charges")
        for label, val in [
            ("Penal", _fmt_num(ch.get("penal", 0))),
            ("Recurring", _fmt_num(ch.get("recurring", 0))),
            ("Non-Recurring", _fmt_num(ch.get("non_recurring", 0))),
            ("Total Charges", _fmt_num(ch.get("total", 0))),
        ]:
            row = _write_kv_row(ws, row, label, val)
        row += 1

    # Section C: Account-wise Interest
    account_interests = data.get("account_interests", [])
    if account_interests:
        row = _write_section_header(ws, row, "C. Account-wise Interest")
        for item in account_interests:
            row = _write_kv_row(ws, row, item["name"], _fmt_num(item["interest"]))
        row += 1

    # WCDL Summary
    wcdl_summary = data.get("wcdl_summary", [])
    if wcdl_summary:
        row = _write_section_header(ws, row, "WCDL Tracker Summary")
        wcdl_headers = ["Loan No.", "Start", "Maturity", "Amount", "ROI", "Status"]
        for ci, h in enumerate(wcdl_headers):
            cell = ws.cell(row=row, column=ci + 1, value=h)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.border = THIN_BORDER
        row += 1
        for loan in wcdl_summary:
            for ci, key in enumerate(["loan_number", "start", "maturity", "principal_str", "roi_str", "status"]):
                ws.cell(row=row, column=ci + 1, value=loan.get(key, "")).font = DATA_FONT
                ws.cell(row=row, column=ci + 1).border = THIN_BORDER
            row += 1
        row += 1


# ── Tab 2: Detail Sheet (fully dynamic columns) ──────────────────────────────

def _build_detail_sheet_tab(wb: Workbook, data: dict):
    ws = wb.create_sheet("Detail Sheet")

    # Dynamic headers: Date + one column per account discovered
    account_columns = data.get("account_columns", [])
    headers = ["Date"] + [c["name"] for c in account_columns]
    widths = [14] + [20] * len(account_columns)

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    daily_rows = data.get("detail_rows", [])
    for i, d in enumerate(daily_rows):
        row = i + 2
        ws.cell(row=row, column=1, value=d.get("date")).number_format = 'DD-MMM-YYYY'

        for col_idx, ac in enumerate(account_columns):
            cell = ws.cell(row=row, column=col_idx + 2, value=d.get(ac["key"], 0))
            cell.number_format = NUM_FMT

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.font = DATA_FONT
            if i % 2 == 0:
                cell.fill = ALT_FILL

    if daily_rows:
        n = len(daily_rows)
        total_row = n + 2
        avg_row = n + 3
        ws.cell(row=total_row, column=1, value="Totals").font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=avg_row, column=1, value="Averages").font = Font(name="Arial", bold=True, size=10)

        for col in range(2, len(headers) + 1):
            letter = get_column_letter(col)
            ws.cell(row=total_row, column=col, value=f"=SUM({letter}2:{letter}{n+1})").number_format = NUM_FMT
            ws.cell(row=avg_row, column=col, value=f"=AVERAGE({letter}2:{letter}{n+1})").number_format = NUM_FMT
            ws.cell(row=total_row, column=col).font = Font(name="Arial", bold=True, size=9)
            ws.cell(row=avg_row, column=col).font = Font(name="Arial", bold=True, size=9)
            ws.cell(row=total_row, column=col).fill = TOTAL_FILL


# ── Tab 3: Forex (only created if forex data exists) ─────────────────────────

def _build_forex_tab(wb: Workbook, data: dict, market_rates: dict):
    ws = wb.create_sheet("Forex")
    headers = [
        "Sr No", "BOE Date", "Value Date", "Drawer Name", "Bill Reference",
        "Currency", "Bill Amount", "Rate-As per Advice", "MKT Avg Rate",
        "Excess vs Avg (₹)", "Excess vs High (₹)",
        "Bill Amt INR (₹)", "Bill Commission (₹)", "Swift Charges (₹)",
        "Total Amt INR (₹)",
    ]
    widths = [8, 14, 14, 28, 20, 10, 18, 14, 14, 18, 18, 20, 16, 16, 20]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36

    from app.computation.forex_excess import forex_excess_vs_avg, forex_excess_vs_high

    forex_txns = data.get("forex_transactions", [])
    for i, txn in enumerate(forex_txns):
        row = i + 2
        fc = txn.get("fc_amount", 0)
        br = txn.get("bank_rate", 0)
        mkt_avg = txn.get("market_avg", 0)
        mkt_high = txn.get("market_high", 0)
        bill_inr = float(fc) * float(br) if br else 0
        comm = txn.get("bill_commission", 0)
        swift = txn.get("swift_charges", 0)

        excess_avg = float(forex_excess_vs_avg(fc, br, mkt_avg)) if mkt_avg else 0
        excess_high = float(forex_excess_vs_high(fc, br, mkt_high)) if mkt_high else 0

        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=txn.get("boe_date", ""))
        ws.cell(row=row, column=3, value=txn.get("value_date", ""))
        ws.cell(row=row, column=4, value=txn.get("drawer_name", ""))
        ws.cell(row=row, column=5, value=txn.get("bill_reference", ""))
        ws.cell(row=row, column=6, value=txn.get("currency", ""))
        ws.cell(row=row, column=7, value=float(fc)).number_format = NUM_FMT
        ws.cell(row=row, column=8, value=float(br)).number_format = '#,##0.0000'
        ws.cell(row=row, column=9, value=float(mkt_avg)).number_format = '#,##0.0000'
        ws.cell(row=row, column=10, value=excess_avg).number_format = NUM_FMT
        ws.cell(row=row, column=11, value=excess_high).number_format = NUM_FMT
        ws.cell(row=row, column=12, value=bill_inr).number_format = NUM_FMT
        ws.cell(row=row, column=13, value=float(comm)).number_format = NUM_FMT
        ws.cell(row=row, column=14, value=float(swift)).number_format = NUM_FMT
        ws.cell(row=row, column=15, value=bill_inr + float(comm) + float(swift)).number_format = NUM_FMT

        for col in range(1, 16):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.font = DATA_FONT
            if i % 2 == 0:
                cell.fill = ALT_FILL

    if forex_txns:
        n = len(forex_txns)
        total_row = n + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
        for col in [7, 10, 11, 12, 13, 14, 15]:
            letter = get_column_letter(col)
            ws.cell(row=total_row, column=col, value=f"=SUM({letter}2:{letter}{n+1})").number_format = NUM_FMT
            ws.cell(row=total_row, column=col).font = Font(name="Arial", bold=True, size=9)
            ws.cell(row=total_row, column=col).fill = TOTAL_FILL


# ── Dynamic Data Extraction from Working Sheet ────────────────────────────────

def _extract_ws_data(ws_wb) -> dict:
    """
    Discover all sheets in the Working Sheet and extract data dynamically.
    No hardcoded account names — everything is derived from what's there.
    """
    data = {
        "month_label": "Report",
        "avg_cc_util_inr": 0,
        "avg_wcdl_util_inr": 0,
        "avg_total_util_inr": 0,
        "total_interest": 0,
        "finance_cost_pct": 0,
        "charges": {"penal": 0, "recurring": 0, "non_recurring": 0, "total": 0},
        "account_interests": [],
        "wcdl_summary": [],
        "forex_transactions": [],
        "has_forex": False,
        "account_columns": [],
        "detail_rows": [],
    }

    # Classify each sheet in the Working Sheet
    account_sheets = []   # sheets with daily balance data
    for name in ws_wb.sheetnames:
        if name in SKIP_SHEETS:
            continue

        name_lower = name.lower()
        if any(kw in name_lower for kw in CC_KEYWORDS):
            account_sheets.append({"name": name, "type": "CC"})
        elif any(kw in name_lower for kw in CURRENT_KEYWORDS):
            account_sheets.append({"name": name, "type": "CURRENT"})
        elif len(name) == 3 and name.upper() == name:
            pass  # currency tab (e.g. "GBP", "USD") — skip for detail sheet
        else:
            account_sheets.append({"name": name, "type": "OTHER"})

    # Build dynamic account columns for Detail Sheet
    account_columns = []
    for sheet_info in account_sheets:
        key = sheet_info["name"].replace(" ", "_").replace("-", "_").lower()
        account_columns.append({"name": sheet_info["name"], "key": key, "type": sheet_info["type"]})
    data["account_columns"] = account_columns

    # Read daily data from each account sheet and build detail rows
    daily_data_per_account = {}
    for ac in account_columns:
        if ac["name"] in ws_wb.sheetnames:
            ws = ws_wb[ac["name"]]
            daily = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    break
                dt = row[0]
                # Balance is in different columns depending on sheet type
                # CC tabs: col D (idx 3) = CC Drawn, Current tabs: col B (idx 1) = Closing Balance
                if ac["type"] == "CC":
                    val = _safe_float(row[3]) if len(row) > 3 else _safe_float(row[1])
                else:
                    val = _safe_float(row[1])
                daily[dt] = val
            daily_data_per_account[ac["key"]] = daily

    # Build detail rows — one row per date, one column per account
    all_dates = set()
    for daily in daily_data_per_account.values():
        all_dates.update(daily.keys())
    sorted_dates = sorted(all_dates)

    detail_rows = []
    for dt in sorted_dates:
        row_data = {"date": dt}
        for ac in account_columns:
            row_data[ac["key"]] = daily_data_per_account.get(ac["key"], {}).get(dt, 0)
        detail_rows.append(row_data)
    data["detail_rows"] = detail_rows

    # Determine month label from first date
    if sorted_dates:
        first_date = sorted_dates[0]
        if isinstance(first_date, datetime.datetime):
            first_date = first_date.date()
        if isinstance(first_date, datetime.date):
            data["month_label"] = first_date.strftime("%b-%y")

    # Extract from Interest tab
    if "Interest" in ws_wb.sheetnames:
        ws = ws_wb["Interest"]
        for row in ws.iter_rows(min_row=1, values_only=True):
            label = str(row[0]) if row[0] else ""
            val = _safe_float(row[1]) if len(row) > 1 else 0

            if "TOTAL FINANCE INTEREST" in label.upper():
                data["total_interest"] = val
            elif label and val > 0 and "total" not in label.lower() and "summary" not in label.lower():
                # Individual account interest entries
                if label not in ("Account", "Loan Number"):
                    data["account_interests"].append({"name": label, "interest": val})

    # Compute CC avg utilisation from CC account columns
    cc_values = []
    for ac in account_columns:
        if ac["type"] == "CC":
            for row_data in detail_rows:
                v = row_data.get(ac["key"], 0)
                if v:
                    cc_values.append(v)
    if cc_values:
        data["avg_cc_util_inr"] = sum(cc_values) / len(cc_values)

    data["avg_total_util_inr"] = data["avg_cc_util_inr"] + data["avg_wcdl_util_inr"]

    if data["avg_total_util_inr"] > 0 and data["total_interest"] > 0:
        from app.computation.finance_cost import finance_cost_pct
        fc = finance_cost_pct(data["total_interest"], data["avg_total_util_inr"])
        data["finance_cost_pct"] = float(fc)

    # Extract from Charges tab
    if "Charges" in ws_wb.sheetnames:
        ws = ws_wb["Charges"]
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row:
                continue
            # Charges tab has: Date | Source | Particulars | Amount | Dr/Cr | Category
            # Summary section has labels in col C (idx 2)
            label = ""
            amt = 0
            if len(row) > 3:
                label = str(row[2]).strip() if row[2] else ""
                amt = _safe_float(row[3])

            if label == "Penal":
                data["charges"]["penal"] = amt
            elif label == "Recurring":
                data["charges"]["recurring"] = amt
            elif label == "Non-Recurring":
                data["charges"]["non_recurring"] = amt
            elif label == "Total":
                data["charges"]["total"] = amt

    # Extract from WCDL Tracker tab
    if "WCDL Tracker" in ws_wb.sheetnames:
        ws = ws_wb["WCDL Tracker"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None or str(row[0]).strip() == "":
                break
            principal = _safe_float(row[4]) if len(row) > 4 else 0
            data["wcdl_summary"].append({
                "loan_number": str(row[0]),
                "start": str(row[1]) if row[1] else "",
                "maturity": str(row[2]) if row[2] else "",
                "principal_str": f"{principal/1e7:.0f} Cr" if principal > 0 else "",
                "roi_str": str(row[5]) if len(row) > 5 and row[5] else "",
                "status": str(row[9]) if len(row) > 9 and row[9] else "",
            })

        # Compute WCDL avg utilisation
        wcdl_values = [_safe_float(r[4]) for r in ws.iter_rows(min_row=2, values_only=True) if r[0] and r[0] != "Total"]
        if wcdl_values:
            data["avg_wcdl_util_inr"] = sum(wcdl_values) / len(wcdl_values) if wcdl_values else 0
            data["avg_total_util_inr"] = data["avg_cc_util_inr"] + data["avg_wcdl_util_inr"]

    # Check for forex
    data["has_forex"] = "Forex" in ws_wb.sheetnames or any(
        name in ws_wb.sheetnames for name in ["GBP", "USD", "EUR", "AUD", "JPY"]
    )

    return data


# ── Helpers ───────────────────────────────────────────────────────────────────

def _write_section_header(ws, row: int, title: str) -> int:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = SECTION_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER
    return row + 1


def _write_kv_row(ws, row: int, label: str, value: str, highlight: bool = False) -> int:
    ws.cell(row=row, column=1, value=label).font = LABEL_FONT
    val_cell = ws.cell(row=row, column=2, value=value)
    val_cell.font = ALERT_FONT if highlight else VALUE_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2).border = THIN_BORDER
    return row + 1


def _fmt_num(val) -> str:
    try:
        return f"₹{float(val):,.2f}"
    except (ValueError, TypeError):
        return "₹0.00"


def _fmt_pct(val) -> str:
    try:
        return f"{float(val)*100:.2f}%"
    except (ValueError, TypeError):
        return "0.00%"


def _safe_float(val) -> float:
    if val is None:
        return 0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0
