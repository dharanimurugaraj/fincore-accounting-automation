"""
Working Sheet Generator — reads the Cumulative Bank Statement Excel
and produces a Working Sheet with DYNAMIC tabs.

The number and names of tabs depend entirely on what sheets exist
in the Statement Excel (which depends on which PDFs were uploaded).

DYNAMIC TAB GENERATION RULES:
  - One "daily balance" tab per account sheet found in Statement Excel
  - One FX tab per currency found in Forex sheet (if any)
  - WCDL Tracker tab only if a WCDL sheet exists
  - Interest tab aggregating all accounts found
  - Charges tab aggregating all charges found
  - EMI tab only if any EMI-category transactions exist
"""

import datetime
import calendar
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, List, Any, Optional

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.computation.cc_interest import (
    cc_daily_interest, cc_monthly_interest, get_cc_roi,
)
from app.computation.wcdl_interest import wcdl_interest, get_wcdl_roi
from app.computation.idle_loss import notional_interest_loss
from app.computation.finance_cost import finance_cost_pct
from app.computation.roi_verification import actual_roi

# ── Style constants ───────────────────────────────────────────────────────────

HDR_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HDR_FILL = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Arial", size=9)
NUM_FMT = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
ALT_FILL = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
TOTAL_FONT = Font(name="Arial", bold=True, size=10)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

# Account types that are treated as "CC" (drawn balance = negative = utilisation)
CC_TYPES = {"CC", "OD"}
# Account types that are "current" (positive balance = idle cash)
CURRENT_TYPES = {"CURRENT", "SAVINGS"}


def generate_working_sheet(
    statement_excel_path: str,
    forex_rates: dict,
    wcdl_loans: list,
    output_path: str,
    account_registry: list[dict] = None,
) -> str:
    """
    Reads Statement Excel, dynamically discovers account sheets,
    applies formulas, and writes the Working Sheet.
    """
    stmt_wb = load_workbook(statement_excel_path, data_only=True)
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    # Discover what's in the Statement Excel
    discovered = _discover_accounts(stmt_wb, account_registry)
    month_dates = _get_month_dates(stmt_wb)

    cc_accounts = {}      # {sheet_name: daily_balances}
    current_accounts = {}  # {sheet_name: daily_balances}
    wcdl_sheet = None
    forex_sheet = None
    all_charges = []
    all_emi_rows = []

    for info in discovered:
        sheet_name = info["sheet_name"]
        acct_type = info["account_type"]

        if acct_type == "WCDL":
            wcdl_sheet = sheet_name
        elif acct_type == "FOREX":
            forex_sheet = sheet_name
        elif acct_type in CC_TYPES:
            daily = _extract_daily_balances(stmt_wb, sheet_name)
            cc_accounts[sheet_name] = daily
            all_charges.extend(_extract_charges(stmt_wb, sheet_name, sheet_name))
            all_emi_rows.extend(_extract_emi_rows(stmt_wb, sheet_name))
        elif acct_type in CURRENT_TYPES:
            daily = _extract_daily_balances(stmt_wb, sheet_name)
            current_accounts[sheet_name] = daily
            all_charges.extend(_extract_charges(stmt_wb, sheet_name, sheet_name))
        else:
            daily = _extract_daily_balances(stmt_wb, sheet_name)
            current_accounts[sheet_name] = daily
            all_charges.extend(_extract_charges(stmt_wb, sheet_name, sheet_name))

    # ── Build tabs dynamically ────────────────────────────────────────────

    # 1. One tab per CC account (with utilisation columns)
    for sheet_name, daily in cc_accounts.items():
        _build_cc_tab(out_wb, daily, sheet_name, wcdl_loans, month_dates)

    # 2. One tab per Current/Savings account
    for sheet_name, daily in current_accounts.items():
        _build_balance_tab(out_wb, daily, sheet_name, month_dates)

    # 3. FX tabs per currency (only if forex data exists)
    if forex_sheet:
        currencies = _extract_forex_currencies(stmt_wb, forex_sheet)
        for currency in currencies:
            _build_fx_tab(out_wb, forex_rates, currency, month_dates)

    # 4. WCDL Tracker (only if WCDL data exists)
    if wcdl_loans:
        _build_wcdl_tracker_tab(out_wb, wcdl_loans, month_dates)

    # 5. Interest tab (aggregates whatever CC accounts + WCDL exist)
    if cc_accounts or wcdl_loans:
        _build_interest_tab(out_wb, cc_accounts, wcdl_loans, month_dates)

    # 6. Charges tab (only if any charges were found)
    if all_charges:
        _build_charges_tab(out_wb, all_charges)

    # 7. EMI tab (only if any EMI transactions were found)
    if all_emi_rows:
        _build_emi_tab(out_wb, all_emi_rows)

    if not out_wb.sheetnames:
        ws = out_wb.create_sheet("No Data")
        ws["A1"] = "No account data found in Statement Excel."

    out_wb.save(output_path)
    return output_path


# ── Account Discovery ─────────────────────────────────────────────────────────

def _discover_accounts(stmt_wb, account_registry: list[dict] = None) -> list[dict]:
    """
    Discover all account sheets in the Statement Excel.
    Uses the account_registry if available, otherwise infers from sheet names.
    """
    if account_registry:
        registry_by_sheet = {r.get("sheet_name"): r for r in account_registry}
    else:
        registry_by_sheet = {}

    discovered = []
    for sheet_name in stmt_wb.sheetnames:
        if sheet_name in ("No Data",):
            continue

        if sheet_name in registry_by_sheet:
            discovered.append(registry_by_sheet[sheet_name])
        else:
            acct_type = _infer_type_from_sheet_name(sheet_name)
            discovered.append({
                "sheet_name": sheet_name,
                "account_type": acct_type,
                "account_id": sheet_name,
                "bank": "Unknown",
                "account_label": sheet_name,
            })

    return discovered


def _infer_type_from_sheet_name(name: str) -> str:
    """Infer account type from a sheet name if no registry is available."""
    name_lower = name.lower()
    if "cc" in name_lower or "cash credit" in name_lower:
        return "CC"
    if "wcdl" in name_lower or "demand loan" in name_lower or "loan" in name_lower:
        return "WCDL"
    if "forex" in name_lower or "remittance" in name_lower:
        return "FOREX"
    if "od" in name_lower or "overdraft" in name_lower:
        return "OD"
    if "savings" in name_lower or "sb " in name_lower:
        return "SAVINGS"
    return "CURRENT"


# ── Tab Builders ──────────────────────────────────────────────────────────────

def _build_cc_tab(wb: Workbook, daily: list, name: str, wcdl_loans: list, month_dates: list):
    """Build a CC/OD account tab with utilisation and WCDL columns."""
    ws = wb.create_sheet(name[:31])
    has_wcdl = bool(wcdl_loans)

    headers = ["Date", "Positive Bal.", "No. Of Days", "CC Drawn"]
    widths = [14, 16, 10, 20]
    if has_wcdl:
        headers.extend(["WCDL", "Total Utilisation"])
        widths.extend([20, 22])

    _write_header_row(ws, headers, widths)

    cc_roi = float(get_cc_roi(month_dates[0] if month_dates else datetime.date.today()))
    wcdl_roi = float(get_wcdl_roi(month_dates[0] if month_dates else datetime.date.today()))

    bal_map = {row["date"]: row["balance"] for row in daily}

    for i, dt in enumerate(month_dates):
        row = i + 2
        raw_bal = bal_map.get(dt, 0)
        cc_bal = abs(raw_bal) if raw_bal < 0 else 0
        positive = raw_bal if raw_bal > 0 else 0

        ws.cell(row=row, column=1, value=dt).number_format = 'DD-MMM-YYYY'
        ws.cell(row=row, column=2, value=positive if positive > 0 else "").number_format = NUM_FMT
        ws.cell(row=row, column=3, value=1)
        ws.cell(row=row, column=4, value=cc_bal).number_format = NUM_FMT

        if has_wcdl:
            wcdl_total = _wcdl_outstanding_on_date(wcdl_loans, dt)
            ws.cell(row=row, column=5, value=wcdl_total).number_format = NUM_FMT
            ws.cell(row=row, column=6, value=cc_bal + wcdl_total).number_format = NUM_FMT

        _style_data_row(ws, row, len(headers), i)

    total_row = len(month_dates) + 2
    _write_total_label(ws, total_row, "Totals")
    sum_cols = [4]
    if has_wcdl:
        sum_cols.extend([5, 6])
    for col in sum_cols:
        _write_column_sum(ws, total_row, col, 2, total_row - 1)

    avg_row = total_row + 1
    _write_total_label(ws, avg_row, "Average")
    for col in sum_cols:
        _write_column_avg(ws, avg_row, col, 2, total_row - 1)


def _build_balance_tab(wb: Workbook, daily: list, name: str, month_dates: list):
    """Build a balance tab for a current/savings account."""
    ws = wb.create_sheet(name[:31])
    headers = ["Date", "Closing Balance", "No. Of Days"]
    widths = [14, 22, 12]
    _write_header_row(ws, headers, widths)

    bal_map = {row["date"]: row["balance"] for row in daily}
    days_positive = 0

    for i, dt in enumerate(month_dates):
        row = i + 2
        bal = bal_map.get(dt, 0)
        if bal > 0:
            days_positive += 1

        ws.cell(row=row, column=1, value=dt).number_format = 'DD-MMM-YYYY'
        ws.cell(row=row, column=2, value=bal).number_format = NUM_FMT
        ws.cell(row=row, column=3, value=1 if bal > 0 else 0)
        _style_data_row(ws, row, 3, i)

    total_row = len(month_dates) + 2
    _write_total_label(ws, total_row, "Total")
    _write_column_sum(ws, total_row, 2, 2, total_row - 1)
    ws.cell(row=total_row, column=3, value=days_positive).font = TOTAL_FONT

    avg_row = total_row + 1
    _write_total_label(ws, avg_row, "Average Balance")
    _write_column_avg(ws, avg_row, 2, 2, total_row - 1)


def _build_fx_tab(wb: Workbook, forex_rates: dict, currency: str, month_dates: list):
    """Build an FX tab for a single currency."""
    ws = wb.create_sheet(currency[:31])
    headers = ["Date", "Rate (INR)", "No. Of Days"]
    widths = [14, 16, 12]
    _write_header_row(ws, headers, widths)

    for i, dt in enumerate(month_dates):
        row = i + 2
        date_str = dt.isoformat()
        rate = forex_rates.get(date_str, {}).get(currency, 0)

        ws.cell(row=row, column=1, value=dt).number_format = 'DD-MMM-YYYY'
        ws.cell(row=row, column=2, value=rate).number_format = '#,##0.0000'
        ws.cell(row=row, column=3, value=1)
        _style_data_row(ws, row, 3, i)

    total_row = len(month_dates) + 2
    _write_total_label(ws, total_row, "Average Rate")
    _write_column_avg(ws, total_row, 2, 2, total_row - 1)


def _build_wcdl_tracker_tab(wb: Workbook, wcdl_loans: list, month_dates: list):
    """Build the WCDL Tracker tab from loan data."""
    ws = wb.create_sheet("WCDL Tracker")
    headers = [
        "Loan Number", "Start Date", "Maturity Date", "Prepayment Date",
        "Principal (₹)", "ROI (%)", "Tenure (Days)",
        "Interest (₹)", "Month Interest (₹)", "Status",
    ]
    widths = [25, 14, 14, 14, 22, 10, 12, 20, 20, 12]
    _write_header_row(ws, headers, widths)

    month_start = month_dates[0] if month_dates else datetime.date(2026, 2, 1)
    month_end = month_dates[-1] if month_dates else datetime.date(2026, 2, 28)

    for i, loan in enumerate(wcdl_loans):
        row = i + 2
        start = _safe_date(loan.get("start_date"))
        maturity = _safe_date(loan.get("maturity_date"))
        prepay = _safe_date(loan.get("prepayment_date"))

        end = prepay if prepay else maturity
        tenure = (end - start).days if start and end else 0

        principal = loan.get("principal", 0)
        roi = loan.get("roi", 0)
        interest_full = float(wcdl_interest(principal, roi, tenure))

        # Compute interest portion within this month only
        if start and end:
            active_from = max(start, month_start)
            active_to = min(end, month_end)
            month_days = max(0, (active_to - active_from).days)
        else:
            month_days = 0
        month_interest = float(wcdl_interest(principal, roi, month_days))

        status = "CLOSED" if prepay else "ACTIVE"

        ws.cell(row=row, column=1, value=loan.get("loan_number", ""))
        ws.cell(row=row, column=2, value=start).number_format = 'DD-MMM-YYYY'
        ws.cell(row=row, column=3, value=maturity).number_format = 'DD-MMM-YYYY'
        ws.cell(row=row, column=4, value=prepay if prepay else "—")
        ws.cell(row=row, column=5, value=float(principal)).number_format = NUM_FMT
        ws.cell(row=row, column=6, value=f"{roi*100:.2f}%")
        ws.cell(row=row, column=7, value=tenure)
        ws.cell(row=row, column=8, value=interest_full).number_format = NUM_FMT
        ws.cell(row=row, column=9, value=month_interest).number_format = NUM_FMT
        ws.cell(row=row, column=10, value=status)
        _style_data_row(ws, row, 10, i)

        status_fill = (
            PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            if status == "ACTIVE"
            else PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        )
        ws.cell(row=row, column=10).fill = status_fill

    total_row = len(wcdl_loans) + 2
    ws.cell(row=total_row, column=7, value="Total").font = TOTAL_FONT
    _write_column_sum(ws, total_row, 8, 2, total_row - 1)
    _write_column_sum(ws, total_row, 9, 2, total_row - 1)


def _build_interest_tab(wb: Workbook, cc_accounts: dict, wcdl_loans: list, month_dates: list):
    """Build the Interest tab — aggregates all CC accounts + WCDL loans found."""
    ws = wb.create_sheet("Interest")

    ref_date = month_dates[0] if month_dates else datetime.date.today()
    cc_roi = float(get_cc_roi(ref_date))

    ws.cell(row=1, column=1, value="CC Interest Summary").font = Font(name="Arial", bold=True, size=11)
    for ci, h in enumerate(["Account", "Monthly Interest (₹)", "ROI (%)"]):
        cell = ws.cell(row=2, column=ci + 1, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12

    cc_total = Decimal("0")
    row = 3
    for name, daily in cc_accounts.items():
        monthly = cc_monthly_interest(daily, cc_roi) if daily else Decimal("0")
        cc_total += monthly
        ws.cell(row=row, column=1, value=name).font = DATA_FONT
        ws.cell(row=row, column=2, value=float(monthly)).number_format = NUM_FMT
        ws.cell(row=row, column=3, value=f"{cc_roi*100:.2f}%")
        _style_data_row(ws, row, 3, row - 3)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Total CC Interest").font = TOTAL_FONT
    ws.cell(row=row, column=2, value=float(cc_total)).number_format = NUM_FMT
    ws.cell(row=row, column=2).font = TOTAL_FONT
    ws.cell(row=row, column=2).fill = TOTAL_FILL
    row += 2

    # WCDL section
    wcdl_total = Decimal("0")
    if wcdl_loans:
        month_start = month_dates[0] if month_dates else datetime.date.today()
        month_end = month_dates[-1] if month_dates else datetime.date.today()

        ws.cell(row=row, column=1, value="WCDL Interest Summary").font = Font(name="Arial", bold=True, size=11)
        row += 1
        for ci, h in enumerate(["Loan Number", "Full Interest (₹)", "Month Interest (₹)", "ROI (%)"]):
            cell = ws.cell(row=row, column=ci + 1, value=h)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
        row += 1

        for idx, loan in enumerate(wcdl_loans):
            start = _safe_date(loan.get("start_date"))
            maturity = _safe_date(loan.get("maturity_date"))
            prepay = _safe_date(loan.get("prepayment_date"))
            end = prepay if prepay else maturity
            tenure = (end - start).days if start and end else 0
            principal = loan.get("principal", 0)
            roi = loan.get("roi", 0)
            interest_full = wcdl_interest(principal, roi, tenure)

            if start and end:
                active_from = max(start, month_start)
                active_to = min(end, month_end)
                month_days = max(0, (active_to - active_from).days)
            else:
                month_days = 0
            month_int = wcdl_interest(principal, roi, month_days)
            wcdl_total += month_int

            ws.cell(row=row, column=1, value=loan.get("loan_number", ""))
            ws.cell(row=row, column=2, value=float(interest_full)).number_format = NUM_FMT
            ws.cell(row=row, column=3, value=float(month_int)).number_format = NUM_FMT
            ws.cell(row=row, column=4, value=f"{roi*100:.2f}%")
            _style_data_row(ws, row, 4, idx)
            row += 1

        ws.cell(row=row, column=1, value="Total WCDL Interest").font = TOTAL_FONT
        ws.cell(row=row, column=3, value=float(wcdl_total)).number_format = NUM_FMT
        ws.cell(row=row, column=3).font = TOTAL_FONT
        ws.cell(row=row, column=3).fill = TOTAL_FILL
        row += 2

    # Grand total
    grand_total = cc_total + wcdl_total
    ws.cell(row=row, column=1, value="TOTAL FINANCE INTEREST").font = Font(name="Arial", bold=True, size=11, color="0D1B2A")
    ws.cell(row=row, column=2, value=float(grand_total)).number_format = NUM_FMT
    ws.cell(row=row, column=2).font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=row, column=2).fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")


def _build_charges_tab(wb: Workbook, all_charges: list):
    """Build the Charges tab from aggregated charges across all accounts."""
    ws = wb.create_sheet("Charges")

    ws.cell(row=1, column=1, value="Bank Charges Summary").font = Font(name="Arial", bold=True, size=11)
    left_headers = ["Date", "Source Account", "Particulars", "Amount (₹)", "Dr/Cr", "Category"]
    for ci, h in enumerate(left_headers):
        cell = ws.cell(row=2, column=ci + 1, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = THIN_BORDER
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 16

    penal = Decimal("0")
    recurring = Decimal("0")
    non_recurring = Decimal("0")

    for i, ch in enumerate(all_charges):
        row = 3 + i
        ws.cell(row=row, column=1, value=ch.get("date", ""))
        ws.cell(row=row, column=2, value=ch.get("source", ""))
        ws.cell(row=row, column=3, value=ch.get("description", ""))
        amt = Decimal(str(ch.get("amount", 0)))
        ws.cell(row=row, column=4, value=float(amt)).number_format = NUM_FMT
        ws.cell(row=row, column=5, value=ch.get("dr_cr", "Dr"))
        cat = ch.get("category", "Recurring")
        ws.cell(row=row, column=6, value=cat)

        if cat == "Penal":
            penal += amt
        elif cat == "Non-Recurring":
            non_recurring += amt
        else:
            recurring += amt

        _style_data_row(ws, row, 6, i)

    summary_row = 3 + len(all_charges) + 1
    ws.cell(row=summary_row, column=1, value="CHARGES SUMMARY").font = Font(name="Arial", bold=True, size=11)
    for label, val in [("Penal", penal), ("Recurring", recurring), ("Non-Recurring", non_recurring), ("Total", penal + recurring + non_recurring)]:
        summary_row += 1
        ws.cell(row=summary_row, column=3, value=label).font = TOTAL_FONT if label == "Total" else DATA_FONT
        ws.cell(row=summary_row, column=4, value=float(val)).number_format = NUM_FMT
        if label == "Total":
            ws.cell(row=summary_row, column=4).font = TOTAL_FONT
            ws.cell(row=summary_row, column=4).fill = TOTAL_FILL


def _build_emi_tab(wb: Workbook, emi_rows: list):
    """Build the EMI tab from discovered EMI transactions."""
    ws = wb.create_sheet("EMI")
    ws.cell(row=1, column=1, value="EMI Details").font = Font(name="Arial", bold=True, size=11)

    headers = ["Date", "Source Account", "Narration", "EMI Amount (₹)"]
    for ci, h in enumerate(headers):
        cell = ws.cell(row=2, column=ci + 1, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = THIN_BORDER
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 18

    total = 0
    for i, emi in enumerate(emi_rows):
        row = 3 + i
        ws.cell(row=row, column=1, value=emi.get("date", "")).font = DATA_FONT
        ws.cell(row=row, column=2, value=emi.get("source", "")).font = DATA_FONT
        ws.cell(row=row, column=3, value=emi.get("narration", "")).font = DATA_FONT
        ws.cell(row=row, column=4, value=emi.get("debit", 0)).number_format = NUM_FMT
        total += emi.get("debit", 0)
        _style_data_row(ws, row, 4, i)

    total_row = 3 + len(emi_rows)
    ws.cell(row=total_row, column=3, value="Total").font = TOTAL_FONT
    ws.cell(row=total_row, column=4, value=total).number_format = NUM_FMT
    ws.cell(row=total_row, column=4).font = TOTAL_FONT
    ws.cell(row=total_row, column=4).fill = TOTAL_FILL


# ── Data Extraction Helpers ───────────────────────────────────────────────────

def _extract_daily_balances(wb, sheet_name: str) -> list:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    balances = []
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        dt = row[0]
        balance = row[5] if len(row) > 5 else 0
        if dt is None:
            continue
        if isinstance(dt, datetime.datetime):
            dt = dt.date()
        elif isinstance(dt, str):
            dt = _safe_date(dt)
        if dt and balance is not None:
            try:
                balances.append({"date": dt, "balance": float(balance)})
            except (ValueError, TypeError):
                continue
    return balances


def _extract_charges(wb, sheet_name: str, source_label: str) -> list:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    charges = []
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        category = row[6] if len(row) > 6 else ""
        if category == "Charges":
            charges.append({
                "date": str(row[0]) if row[0] else "",
                "source": source_label,
                "description": row[1] or "",
                "amount": float(row[3]) if row[3] else float(row[4]) if row[4] else 0,
                "dr_cr": "Dr" if row[3] else "Cr",
                "category": "Recurring",
            })
    return charges


def _extract_emi_rows(wb, sheet_name: str) -> list:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    emis = []
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        category = row[6] if len(row) > 6 else ""
        if category == "EMI":
            emis.append({
                "date": str(row[0]) if row[0] else "",
                "source": sheet_name,
                "narration": row[1] or "",
                "debit": float(row[3]) if row[3] else 0,
            })
    return emis


def _extract_forex_currencies(wb, sheet_name: str) -> list:
    """Get unique currencies from the Forex sheet."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    currencies = set()
    for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
        currency = row[5] if len(row) > 5 else None
        if currency and isinstance(currency, str) and len(currency) == 3:
            currencies.add(currency.upper())
    return sorted(currencies)


def _get_month_dates(wb) -> list:
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            dt = row[0]
            if dt is None:
                continue
            if isinstance(dt, datetime.datetime):
                dt = dt.date()
            elif isinstance(dt, str):
                dt = _safe_date(dt)
            if dt:
                year, month = dt.year, dt.month
                last_day = calendar.monthrange(year, month)[1]
                return [datetime.date(year, month, d) for d in range(1, last_day + 1)]
    return [datetime.date.today().replace(day=d) for d in range(1, 29)]


def _wcdl_outstanding_on_date(loans: list, dt: datetime.date) -> float:
    total = 0
    for loan in loans:
        start = _safe_date(loan.get("start_date"))
        maturity = _safe_date(loan.get("maturity_date"))
        prepay = _safe_date(loan.get("prepayment_date"))
        end = prepay if prepay else maturity
        if start and end and start <= dt <= end:
            total += loan.get("principal", 0)
    return float(total)


# ── Formatting Helpers ────────────────────────────────────────────────────────

def _write_header_row(ws, headers: list, widths: list):
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _style_data_row(ws, row: int, max_col: int, row_index: int):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        if not cell.font or cell.font == Font():
            cell.font = DATA_FONT
        if row_index % 2 == 0 and cell.fill == PatternFill():
            cell.fill = ALT_FILL


def _write_total_label(ws, row: int, label: str):
    ws.cell(row=row, column=1, value=label).font = TOTAL_FONT
    ws.cell(row=row, column=1).fill = TOTAL_FILL


def _write_column_sum(ws, total_row: int, col: int, start_row: int, end_row: int):
    letter = get_column_letter(col)
    cell = ws.cell(row=total_row, column=col)
    cell.value = f"=SUM({letter}{start_row}:{letter}{end_row})"
    cell.number_format = NUM_FMT
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_FILL


def _write_column_avg(ws, avg_row: int, col: int, start_row: int, end_row: int):
    letter = get_column_letter(col)
    cell = ws.cell(row=avg_row, column=col)
    cell.value = f"=AVERAGE({letter}{start_row}:{letter}{end_row})"
    cell.number_format = NUM_FMT
    cell.font = TOTAL_FONT


def _safe_date(val) -> Optional[datetime.date]:
    if val is None:
        return None
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, str):
        val = val.strip()
        if not val or val == "—":
            return None
        try:
            parts = val.split("-")
            if len(parts) == 3:
                return datetime.date(int(parts[0]), int(parts[1]), int(parts[2]))
        except (ValueError, IndexError):
            pass
    return None
