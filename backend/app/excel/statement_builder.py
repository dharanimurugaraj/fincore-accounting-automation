"""
Builds the Cumulative Bank Statement Excel from OCR-extracted JSON data.

OUTPUT: One sheet per uploaded account (dynamic — not fixed).
The number and names of sheets depend entirely on the PDFs that were uploaded
and classified. Each account gets its own sheet named from the account_label.

Column schema per account sheet: Date | Narration | Reference | Debit | Credit | Balance | Category
WCDL sheets: Loan Number | Start | Maturity | Prepay | Principal | ROI | Tenure | Interest | Status
Forex sheets: Sr | BOE Date | Value Date | Drawer | Bill Ref | Currency | FC Amt | Bank Rate | RBI Rate | INR | Commission | Swift | Confidence
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, Any, List, Optional
import datetime

# ── Style Constants ───────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DEBIT_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
DEBIT_FONT = Font(name="Arial", size=9, color="C00000")

CREDIT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
CREDIT_FONT = Font(name="Arial", size=9, color="375623")

BALANCE_NEG_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
BALANCE_POS_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

CHARGES_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
INTEREST_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
EMI_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")

DATA_FONT = Font(name="Arial", size=9)
DATA_ALIGN = Alignment(vertical="center")
NUMBER_FORMAT = '#,##0.00'
DATE_FORMAT = 'DD-MMM-YYYY'

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

HEADERS = ["Date", "Narration", "Reference No.", "Debit (₹)", "Credit (₹)", "Balance (₹)", "Category"]
COL_WIDTHS = [14, 45, 20, 18, 18, 20, 15]


def build_statement_excel(
    extracted_data: Dict[str, Any],
    forex_rates: Dict[str, Dict[str, float]],
    output_path: str,
    account_registry: list[dict] = None,
) -> str:
    """
    Build the Cumulative Bank Statement Excel — one sheet per account.

    Args:
        extracted_data: {account_id: ocr_output_json} — keyed by whatever IDs
                        the classifier assigned, not hardcoded.
        forex_rates: {date_str: {currency: rate}}
        output_path: Where to save the Excel file.
        account_registry: List of AccountMeta dicts from the classifier, used
                         to determine sheet names and account types.
    """
    wb = Workbook()
    wb.remove(wb.active)

    if not account_registry:
        account_registry = _build_registry_from_data(extracted_data)

    registry_by_id = {r["account_id"]: r for r in account_registry}

    for account_id, data in extracted_data.items():
        meta = registry_by_id.get(account_id, {})
        acct_type = meta.get("account_type", _infer_type(data))
        sheet_name = meta.get("sheet_name", f"Account {account_id}")[:31]

        # Prevent duplicate sheet names
        sheet_name = _unique_sheet_name(wb, sheet_name)

        ws = wb.create_sheet(sheet_name)

        if acct_type == "WCDL":
            _build_wcdl_sheet(ws, data)
        elif acct_type == "FOREX":
            _build_forex_sheet(ws, data, forex_rates)
        else:
            _build_account_sheet(ws, data, account_id)

    if not wb.sheetnames:
        ws = wb.create_sheet("No Data")
        ws["A1"] = "No account data available for this period."

    wb.save(output_path)
    return output_path


def _build_registry_from_data(extracted_data: dict) -> list[dict]:
    """Build a minimal account registry if none was provided (backward compat)."""
    registry = []
    for acct_id, data in extracted_data.items():
        acct_type = _infer_type(data)
        registry.append({
            "account_id": acct_id,
            "account_type": acct_type,
            "bank": "Unknown",
            "sheet_name": f"{acct_id} {acct_type} AC"[:31],
            "account_label": f"{acct_id} {acct_type}",
        })
    return registry


def _infer_type(data: dict) -> str:
    """Infer account type from the shape of the OCR output data."""
    if "loans" in data:
        return "WCDL"
    if data.get("transactions") and any("boe_date" in t for t in data["transactions"][:3]):
        return "FOREX"
    return "CC"


def _unique_sheet_name(wb: Workbook, name: str) -> str:
    """Ensure sheet name is unique by appending a number if needed."""
    if name not in wb.sheetnames:
        return name
    for i in range(2, 100):
        candidate = f"{name[:28]}({i})"
        if candidate not in wb.sheetnames:
            return candidate
    return name[:28] + "(x)"


# ── Account Sheet Builder (CC, Current, Savings, OD — any transactional account)

def _build_account_sheet(ws, data: dict, account_id: str):
    """Build a standard account transaction sheet with formatting and validation."""
    _write_headers(ws)

    transactions = data.get("transactions", [])
    opening_balance = data.get("opening_balance", 0)
    closing_balance = data.get("closing_balance", 0)

    total_debits = Decimal("0")
    total_credits = Decimal("0")
    row_num = 2

    for i, txn in enumerate(transactions):
        row_num = i + 2
        is_odd = (i % 2 == 0)

        dt = _parse_date(txn.get("date", ""))
        narration = txn.get("narration", "")
        reference = txn.get("reference")
        debit = txn.get("debit")
        credit = txn.get("credit")
        balance = txn.get("balance", 0)
        category = txn.get("category", "")

        ws.cell(row=row_num, column=1, value=dt).number_format = DATE_FORMAT
        ws.cell(row=row_num, column=2, value=narration)
        ws.cell(row=row_num, column=3, value=reference or "")
        ws.cell(row=row_num, column=7, value=category)

        debit_cell = ws.cell(row=row_num, column=4)
        if debit and float(debit) > 0:
            debit_cell.value = float(debit)
            debit_cell.number_format = NUMBER_FORMAT
            debit_cell.font = DEBIT_FONT
            debit_cell.fill = DEBIT_FILL
            total_debits += Decimal(str(debit))

        credit_cell = ws.cell(row=row_num, column=5)
        if credit and float(credit) > 0:
            credit_cell.value = float(credit)
            credit_cell.number_format = NUMBER_FORMAT
            credit_cell.font = CREDIT_FONT
            credit_cell.fill = CREDIT_FILL
            total_credits += Decimal(str(credit))

        bal_cell = ws.cell(row=row_num, column=6)
        bal_cell.value = float(balance)
        bal_cell.number_format = NUMBER_FORMAT
        if float(balance) < 0:
            bal_cell.fill = BALANCE_NEG_FILL
            bal_cell.font = DEBIT_FONT
        else:
            bal_cell.fill = BALANCE_POS_FILL
            bal_cell.font = CREDIT_FONT

        row_fill = None
        if category == "Charges":
            row_fill = CHARGES_FILL
        elif category == "Interest":
            row_fill = INTEREST_FILL
        elif category == "EMI":
            row_fill = EMI_FILL
        elif is_odd:
            row_fill = ALT_ROW_FILL

        if row_fill:
            for col in range(1, 8):
                cell = ws.cell(row=row_num, column=col)
                if cell.fill == PatternFill():
                    cell.fill = row_fill

        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.border = THIN_BORDER
            cell.alignment = DATA_ALIGN
            if not cell.font or cell.font == Font():
                cell.font = DATA_FONT

    summary_start = row_num + 2
    _write_summary(ws, summary_start, opening_balance, closing_balance,
                   total_debits, total_credits, len(transactions))

    expected_closing = Decimal(str(opening_balance)) - total_debits + total_credits
    actual_closing = Decimal(str(closing_balance))
    diff = abs(expected_closing - actual_closing)

    if diff > Decimal("1"):
        fail_cell = ws.cell(row=1, column=8)
        fail_cell.value = "BALANCE CHECK FAILED"
        fail_cell.font = Font(name="Arial", bold=True, size=11, color="C00000")
        fail_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        note_cell = ws.cell(row=2, column=8)
        note_cell.value = f"Expected: {expected_closing:.2f} | Actual: {actual_closing:.2f} | Diff: {diff:.2f}"
        note_cell.font = Font(name="Arial", size=9, color="C00000")


# ── WCDL Sheet Builder ────────────────────────────────────────────────────────

def _build_wcdl_sheet(ws, data: dict):
    """Build WCDL Loan Account sheet."""
    headers = [
        "Loan Number", "Start Date", "Maturity Date", "Prepayment Date",
        "Principal (₹)", "ROI (%)", "Tenure (Days)", "Interest (₹)", "Status"
    ]
    widths = [25, 14, 14, 14, 20, 10, 12, 20, 12]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    loans = data.get("loans", [])
    for i, loan in enumerate(loans):
        row = i + 2
        start = _parse_date(loan.get("start_date", ""))
        maturity = _parse_date(loan.get("maturity_date", ""))
        prepay = _parse_date(loan.get("prepayment_date")) if loan.get("prepayment_date") else None

        tenure = 0
        if start and maturity:
            end = prepay if prepay else maturity
            tenure = (end - start).days if isinstance(end, datetime.date) and isinstance(start, datetime.date) else 0

        principal = loan.get("principal", 0)
        roi = loan.get("roi", 0)

        from app.computation.wcdl_interest import wcdl_interest
        interest = float(wcdl_interest(principal, roi, tenure))

        status = "CLOSED" if prepay else "ACTIVE"

        ws.cell(row=row, column=1, value=loan.get("loan_number", ""))
        ws.cell(row=row, column=2, value=start).number_format = DATE_FORMAT
        ws.cell(row=row, column=3, value=maturity).number_format = DATE_FORMAT
        ws.cell(row=row, column=4, value=prepay or "").number_format = DATE_FORMAT if prepay else "@"
        ws.cell(row=row, column=5, value=float(principal)).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=6, value=f"{roi*100:.2f}%")
        ws.cell(row=row, column=7, value=tenure)
        ws.cell(row=row, column=8, value=interest).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=9, value=status)

        fill = ALT_ROW_FILL if i % 2 == 0 else PatternFill()
        status_fill = CREDIT_FILL if status == "ACTIVE" else CHARGES_FILL
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = DATA_ALIGN
            if col == 9:
                cell.fill = status_fill
            elif fill:
                cell.fill = fill


# ── Forex Sheet Builder ───────────────────────────────────────────────────────

def _build_forex_sheet(ws, data: dict, forex_rates: dict):
    """Build Forex Outward Remittance sheet."""
    headers = [
        "Sr No", "BOE Date", "Value Date", "Drawer Name", "Bill Reference",
        "Currency", "FC Amount", "Bank Rate", "RBI Rate", "INR Amount",
        "Bill Commission", "Swift Charges", "Confidence"
    ]
    widths = [8, 14, 14, 30, 20, 10, 18, 12, 12, 20, 15, 15, 12]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    transactions = data.get("transactions", [])
    for i, txn in enumerate(transactions):
        row = i + 2

        boe_date = _parse_date(txn.get("boe_date", ""))
        value_date = _parse_date(txn.get("value_date", ""))
        currency = txn.get("currency", "")
        fc_amount = txn.get("fc_amount", 0)
        bank_rate = txn.get("bank_rate", 0)

        rbi_rate = 0
        vd_str = txn.get("value_date", "")
        if vd_str in forex_rates and currency in forex_rates.get(vd_str, {}):
            rbi_rate = forex_rates[vd_str][currency]

        inr_amount = float(fc_amount) * float(bank_rate) if bank_rate else 0

        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=boe_date).number_format = DATE_FORMAT
        ws.cell(row=row, column=3, value=value_date).number_format = DATE_FORMAT
        ws.cell(row=row, column=4, value=txn.get("drawer_name", ""))
        ws.cell(row=row, column=5, value=txn.get("bill_reference", ""))
        ws.cell(row=row, column=6, value=currency)
        ws.cell(row=row, column=7, value=float(fc_amount)).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=8, value=float(bank_rate)).number_format = '#,##0.0000'
        ws.cell(row=row, column=9, value=float(rbi_rate)).number_format = '#,##0.0000'
        ws.cell(row=row, column=10, value=inr_amount).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=11, value=float(txn.get("bill_commission", 0))).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=12, value=float(txn.get("swift_charges", 0))).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=13, value=txn.get("confidence", 0))

        fill = ALT_ROW_FILL if i % 2 == 0 else PatternFill()
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = DATA_ALIGN
            if fill:
                cell.fill = fill


# ── Helpers ───────────────────────────────────────────────────────────────────

def _write_headers(ws):
    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = "A1:G1"
    ws.freeze_panes = "A2"


def _write_summary(ws, start_row: int, opening: float, closing: float,
                   total_debits: Decimal, total_credits: Decimal, txn_count: int):
    summary_font = Font(name="Arial", bold=True, size=10)
    value_font = Font(name="Arial", size=10)

    labels = [
        ("Monthly Summary", None),
        ("Opening Balance", float(opening)),
        ("Total Debits", float(total_debits)),
        ("Total Credits", float(total_credits)),
        ("Net Movement", float(total_credits - total_debits)),
        ("Closing Balance", float(closing)),
        ("No. of Transactions", txn_count),
    ]
    for i, (label, value) in enumerate(labels):
        row = start_row + i
        ws.cell(row=row, column=2, value=label).font = summary_font
        if value is not None:
            val_cell = ws.cell(row=row, column=4, value=value)
            val_cell.font = value_font
            val_cell.number_format = NUMBER_FORMAT if isinstance(value, float) else "0"
            if isinstance(value, float):
                val_cell.font = Font(name="Arial", size=10, color="C00000" if value < 0 else "375623")


def _parse_date(date_str: Optional[str]) -> Optional[datetime.date]:
    if not date_str:
        return None
    try:
        parts = date_str.split("-")
        if len(parts) == 3:
            return datetime.date(int(parts[0]), int(parts[1]), int(parts[2]))
    except (ValueError, IndexError):
        pass
    return None
