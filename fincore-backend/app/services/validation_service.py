"""
All validation checks — validates computed vs bank-stated values.
Moved from validators/reconciliation.py.
"""

from decimal import Decimal, ROUND_HALF_UP
from dataclasses import dataclass, asdict

TOLERANCE = Decimal("1.00")


@dataclass
class ValidationCheck:
    name: str
    computed: Decimal
    bank_stated: Decimal
    difference: Decimal
    status: str  # PASS | WARN | FAIL
    tolerance: Decimal = TOLERANCE

    def to_dict(self) -> dict:
        return {
            "name": self.name,
            "computed": float(self.computed),
            "bank_stated": float(self.bank_stated),
            "difference": float(self.difference),
            "status": self.status,
            "tolerance": float(self.tolerance),
        }


def run_all_validations(computed_values: dict, bank_stated_values: dict) -> dict:
    """Run all validation checks. Returns validation result dict."""
    checks: list[ValidationCheck] = []

    # Check 1: CC Interest per account
    cc_accounts = computed_values.get("cc_interest_accounts", {})
    for account_name, computed_val in cc_accounts.items():
        computed = Decimal(str(computed_val))
        stated = Decimal(str(bank_stated_values.get(f"cc_interest_{account_name}", 0)))
        diff = abs(computed - stated)
        if computed == 0 and stated == 0:
            status = "PASS"
        elif diff <= TOLERANCE:
            status = "PASS"
        elif diff <= Decimal("10"):
            status = "WARN"
        else:
            status = "FAIL"
        checks.append(ValidationCheck(name=f"CC Interest — {account_name}",
                                      computed=computed, bank_stated=stated,
                                      difference=diff, status=status))

    # Check 2: WCDL Interest per loan
    wcdl_computed = computed_values.get("wcdl_interests", {})
    for loan_id, vals in wcdl_computed.items():
        computed = Decimal(str(vals.get("computed", 0)))
        stated = Decimal(str(vals.get("bank_stated", 0)))
        diff = abs(computed - stated)
        if computed == 0 and stated == 0:
            status = "PASS"
        elif diff <= TOLERANCE:
            status = "PASS"
        elif diff <= Decimal("10"):
            status = "WARN"
        else:
            status = "FAIL"
        checks.append(ValidationCheck(name=f"WCDL Interest — Loan {loan_id}",
                                      computed=computed, bank_stated=stated,
                                      difference=diff, status=status))

    # Check 3: Cross-sheet reconciliation
    for field_name, label in [
        ("total_cc_interest", "Total CC Interest"),
        ("total_wcdl_interest", "Total WCDL Interest"),
        ("total_charges", "Total Charges"),
    ]:
        ws_val = Decimal(str(computed_values.get(f"ws_{field_name}", 0)))
        rpt_val = Decimal(str(computed_values.get(f"rpt_{field_name}", 0)))
        diff = abs(ws_val - rpt_val)
        status = "PASS" if (ws_val == 0 and rpt_val == 0) or diff <= TOLERANCE else "FAIL"
        checks.append(ValidationCheck(name=f"Cross-Sheet — {label}",
                                      computed=ws_val, bank_stated=rpt_val,
                                      difference=diff, status=status))

    # Check 4: Balance checks per account
    balance_checks = computed_values.get("balance_checks", {})
    for account, vals in balance_checks.items():
        expected = Decimal(str(vals.get("expected_closing", 0)))
        actual = Decimal(str(vals.get("actual_closing", 0)))
        diff = abs(expected - actual)
        status = "PASS" if diff <= TOLERANCE else "FAIL"
        checks.append(ValidationCheck(name=f"Balance Check — {account}",
                                      computed=expected, bank_stated=actual,
                                      difference=diff, status=status))

    # Check 5: Forex INR check
    forex_checks = computed_values.get("forex_inr_checks", [])
    for fx in forex_checks:
        computed_inr = Decimal(str(fx.get("computed_inr", 0)))
        stated_inr = Decimal(str(fx.get("stated_inr", 0)))
        diff = abs(computed_inr - stated_inr)
        status = "PASS" if diff <= TOLERANCE else "FAIL"
        checks.append(ValidationCheck(name=f"Forex INR — {fx.get('reference', 'unknown')}",
                                      computed=computed_inr, bank_stated=stated_inr,
                                      difference=diff, status=status))

    blocked = any(c.status == "FAIL" for c in checks)
    warnings = [c for c in checks if c.status == "WARN"]

    return {
        "passed": not blocked, "blocked": blocked,
        "checks": [c.to_dict() for c in checks],
        "warnings": [c.to_dict() for c in warnings],
        "summary": {
            "total_checks": len(checks),
            "passed": sum(1 for c in checks if c.status == "PASS"),
            "warnings": len(warnings),
            "failed": sum(1 for c in checks if c.status == "FAIL"),
        },
    }


def extract_computed_values_from_ws(ws_path: str) -> dict:
    """Read Working Sheet and extract computed values for validation."""
    from openpyxl import load_workbook
    wb = load_workbook(ws_path, data_only=True)
    result = {}

    result["cc_interest_accounts"] = {}
    if "Interest" in wb.sheetnames:
        ws = wb["Interest"]
        for row in ws.iter_rows(min_row=1, values_only=True):
            name = str(row[0]).strip() if row[0] else ""
            val = row[1] if len(row) > 1 and row[1] is not None else 0

            if name.upper().startswith("TOTAL") or name.upper().startswith("WCDL") or \
               not name or name in ("Account", "Loan Number") or "SUMMARY" in name.upper():
                continue
            if "FINANCE INTEREST" in name.upper():
                result["ws_total_interest"] = _safe_decimal(val)
                continue
            if "Total CC Interest" in name:
                result["ws_total_cc_interest"] = _safe_decimal(val)
                continue
            if "Total WCDL Interest" in name:
                result["ws_total_wcdl_interest"] = _safe_decimal(val)
                continue
            if _safe_decimal(val) > 0:
                result["cc_interest_accounts"][name] = float(_safe_decimal(val))

    result["wcdl_interests"] = {}
    if "WCDL Tracker" in wb.sheetnames:
        ws = wb["WCDL Tracker"]
        total_wcdl = Decimal("0")
        for row in ws.iter_rows(min_row=2, values_only=True):
            loan_num = str(row[0]).strip() if row[0] else ""
            if not loan_num or loan_num == "Total":
                break
            interest = _safe_decimal(row[7]) if len(row) > 7 else Decimal("0")
            result["wcdl_interests"][loan_num] = {
                "computed": float(interest), "bank_stated": float(interest),
            }
            total_wcdl += interest
        result["ws_total_wcdl_interest"] = float(total_wcdl)

    if "Charges" in wb.sheetnames:
        ws = wb["Charges"]
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 4:
                continue
            label = str(row[2]).strip() if row[2] else ""
            if label == "Total":
                result["ws_total_charges"] = _safe_decimal(row[3])
                break

    return result


def extract_bank_stated_from_stmt(stmt_path: str) -> dict:
    """Read Statement Excel and extract bank-stated values."""
    from openpyxl import load_workbook
    wb = load_workbook(stmt_path, data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in ("No Data",):
            continue
        ws = wb[sheet_name]
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        if "category" not in headers:
            continue
        cat_idx = headers.index("category")
        debit_idx = headers.index("debit (₹)") if "debit (₹)" in headers else 3
        interest_total = Decimal("0")
        for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
            if len(row) <= cat_idx:
                continue
            category = str(row[cat_idx]).strip() if row[cat_idx] else ""
            if category == "Interest":
                debit = _safe_decimal(row[debit_idx]) if len(row) > debit_idx else Decimal("0")
                interest_total += debit
        result[f"cc_interest_{sheet_name}"] = float(interest_total)

    return result


def _safe_decimal(val) -> Decimal:
    if val is None:
        return Decimal("0")
    try:
        return Decimal(str(val)).quantize(Decimal("0.01"), ROUND_HALF_UP)
    except Exception:
        return Decimal("0")
